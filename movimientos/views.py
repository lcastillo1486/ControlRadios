from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from ordenes.models import ordenRegistro
from movimientos.models import salidasDetalle, contable, abono_factura,vista_ordenes_cxc, controlrxevent
from .forms import radiotipos, agregarInven, formBuscaRadio, guardaEntradaRx, formEntradaDetalle, formBuscarInformes, FacturaPDFForm, formRegistroMontoFact, formRegistroMontopago, comprobantePagoForm, comprobanteabonoForm, formRegistroMontoFactNoSunat, rxcontroleventoform, ResponsableForm, rxcontroleventoformRecojo, FormPedidoCliente
from django.contrib import messages
from .models import movimientoRadios, invSeriales, entradaDetalle, accesoriosFaltantes, radiosFantantes, vista_radios_faltantes, vista_accesorios_faltantes, vista_movimiento_radios_tipos, auditoria, mochila, vista_ordenes_procesadas, vista_ordenes_cerradas, vista_entrada_detalle, vista_movimiento_radios_tipos, CajasMikrot, controlinventario, espejo_inventario_ant, espejo_inventario_desp, inv_accesorios, entrada_salida_acce, controlinventarioacce, espejo_inventarioacce_ant, espejo_inventarioacce_desp, entrada_accesorios, inv_accesorios_temp, rpt_kardex,formulario_pedido, controlrx_event_dia, espejo_dia_control_rx, pedidos_asignados_prepa, pedidos_asignados_entrega
from cliente.models import cliente
from django import forms
from django.db import models
from io import BytesIO
import datetime
from django.utils import timezone 

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, legal, portrait
from reportlab.lib.units import cm
import os
from django.contrib.auth.decorators import login_required
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFont, registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import fonts
from datetime import date, timedelta
from django.db.models.functions import TruncDate
from django.db.models import Count, Sum
from django.db.models import F
from django.db.models.functions import TruncMonth
from django.db.models.functions import ExtractMonth
from django.db.models import Case, When, Value, IntegerField, CharField
from django.db.models.functions import Extract
from collections import defaultdict
import calendar
import matplotlib.pyplot as ptl
import base64
from django.db.models import Q
from datetime import datetime as d
import re
import requests
from requests.auth import HTTPBasicAuth
from pythonping import ping
from django.http import JsonResponse
import json
import openpyxl
from openpyxl.styles import PatternFill, alignment, Alignment
from django.core.validators import RegexValidator
import urllib.parse
from django.contrib.auth.models import User

# Create your views here. 
@login_required
def entradaRadios(request, id, orden_id):

    form = guardaEntradaRx()
    formEntrada = formEntradaDetalle()
    ordenes = ordenRegistro.objects.get(id = orden_id)
    msalida = salidasDetalle.objects.get(id_orden=orden_id)
    #verificar aqui el id
    radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")


    if request.method == 'POST':

        form = guardaEntradaRx(request.POST)
        
        if form.is_valid():
            a = form.cleaned_data['serial']
            ##comprobar que existe una salida con ese serial y ese numero de salida
            if not movimientoRadios.objects.filter(serial = a, id_salida = id, estado = "F").exists():
                initial_data = {
                    'serial': '',
                    }
                form = guardaEntradaRx(initial=initial_data)
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                                        "error": "El serial ingresado no corresponde a la salida", "formEntrada":formEntrada})
            ##valida si ya fue agregado
            if movimientoRadios.objects.filter(serial = a, id_salida = id, estado = "D").exists():
                initial_data = {
                    'serial': '',
                    }
                form = guardaEntradaRx(initial=initial_data)
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                                        "error": "El serial ingresado ya fue agregado", "formEntrada":formEntrada})
            
            else:
                valExisteRxSalida = movimientoRadios.objects.get(serial = a, id_salida = id, estado = "F")
                a = form.save(commit=False)
                a.id_salida = id
                a.estado = 'D'
                a.id_tipo = valExisteRxSalida.id_tipo
                a.save()

                d = a.serial

                #borrar si existe en faltantes

                if radiosFantantes.objects.filter(fserial = d, id_salida = id).exists():
                    borrarrx = radiosFantantes.objects.filter(fserial = d, id_salida = id)
                    borrarrx.delete()


                #cambiar el estatus del radio en el inventario

                cambiaestado =  invSeriales.objects.get(codigo = d)
                cambiaestado.estado_id = 4
                cambiaestado.save()

                ## afectar inventario accesorios

                # BATERIAS

                t_suma_bat = inv_accesorios.objects.get(id = 5)
                t_mas_baterias = 1
                t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias
                t_suma_bat.save()

                fecha_mov = datetime.date.today()

                f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_bat.save()

                


            ordenes = ordenRegistro.objects.get(id = orden_id)
            msalida = salidasDetalle.objects.get(id_orden=orden_id)

            initial_data = {
                    'serial': '',
                    }
            form = guardaEntradaRx(initial=initial_data)

            radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")
            cuenta_radios_cargadas  = movimientoRadios.objects.filter(id_salida = id, estado = "D").count()
            return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                                    "formEntrada":formEntrada, 'cantidad_rx':cuenta_radios_cargadas}) 
@login_required
def entradas(request, id):

    form = guardaEntradaRx()
    formEntrada = formEntradaDetalle()
    context = {'form':form}

    
    msalida = salidasDetalle.objects.get(id=id)
    orden_id = msalida.id_orden
    ordenes = ordenRegistro.objects.get(id = orden_id)

#buscar las radios cargadas

    radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")
    cuenta_radios_cargadas  = movimientoRadios.objects.filter(id_salida = id, estado = "D").count()
    return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, 'cantidad_rx':cuenta_radios_cargadas})
@login_required
def entradatotal(request, id, id_orden):

    if not request.user.is_superuser:
        return redirect('/ordenesCerradas/')

#cambiar el estatus en el inventario de radios 

    radios_inv = movimientoRadios.objects.filter(id_salida = id, estado = "F")
    for i in radios_inv:
        serial = i.serial
        cambiaestado =  invSeriales.objects.get(codigo = serial)
        cambiaestado.estado_id = 4
        cambiaestado.save()

#buscar en movimientosradios todas las radios cargadas y crear un registro nueno con D

    radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "F")
    for i in radiosCargadas:
        salida = i.id_salida
        serial = i.serial
        estado = 'D'
        id_tipo = i.id_tipo_id

        graba_radio = movimientoRadios(id_salida = salida, serial = serial, estado = estado, id_tipo_id = id_tipo)
        graba_radio.save()

# cambiar el estado de la orden 

    cambiaestadoorden = ordenRegistro.objects.filter(id = id_orden).update(estado_id = 4)

#Buscar el nombre del cliente 

    client = ordenRegistro.objects.get(id=id_orden)
    client_id = client.cliente

#busca en movimientosalidadetalle con el ID_ORDEN
    
    salida = salidasDetalle.objects.get(id_orden = id_orden)

#agrega un registro en entrada detalle  
    mov_ent_detalle = entradaDetalle(id_salida = id, id_orden = salida.id_orden, fecha_creacion = salida.fecha_creacion, cobras = salida.cobras, 
                                     cargadores = salida.cargadores, handsfree = salida.handsfree, cascos = salida.cascos, repetidoras = salida.repetidoras,
                                     estaciones = salida.estaciones, baterias = salida.baterias, observaciones = 'Devolución Total', cliente = client_id)
    
    mov_ent_detalle.save()

    ##### ENTRADA DE ACCESORIOS ######

    cuentaRxOrden = movimientoRadios.objects.filter(id_salida = id, estado = "F").count()

    mas_cobras = salida.cobras
    mas_handsfree = salida.handsfree
    mas_escolta = salida.estaciones
    mas_cascos = salida.cascos
    mas_baterias = salida.baterias
    mas_cargadores = salida.cargadores

    fecha_mov = datetime.date.today()

    # COBRAS 
    t_suma_cobras = inv_accesorios.objects.get(id = 1)
    t_suma_cobras.cantidad = t_suma_cobras.cantidad + mas_cobras
    t_suma_cobras.save()

    f_suma_cobras = entrada_salida_acce(id_item = 1, cantidad = mas_cobras,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_cobras.save()

    # HANDSFREE 

    t_suma_hands = inv_accesorios.objects.get (id = 2)
    t_suma_hands.cantidad = t_suma_hands.cantidad + mas_handsfree
    t_suma_hands.save()

    f_suma_hands = entrada_salida_acce(id_item = 2, cantidad = mas_handsfree,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_hands.save()

    # ESCOLTA 

    t_suma_escolta = inv_accesorios.objects.get(id = 3)
    t_suma_escolta.cantidad = t_suma_escolta.cantidad + mas_escolta
    t_suma_escolta.save()

    f_suma_escolta = entrada_salida_acce(id_item = 3, cantidad = mas_escolta,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_escolta.save() 

    # CASCOS

    t_suma_cascos = inv_accesorios.objects.get(id = 4)
    t_suma_cascos.cantidad = t_suma_cascos.cantidad + mas_cascos
    t_suma_cascos.save()

    f_suma_cascos = entrada_salida_acce(id_item = 4, cantidad = mas_cascos,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_cascos.save()

    # BATERIAS

    t_suma_bat = inv_accesorios.objects.get(id = 5)
    t_mas_baterias = mas_baterias + cuentaRxOrden
    t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias
    t_suma_bat.save()

    f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_bat.save()

    # CARGADORES 

    t_suma_carg = inv_accesorios.objects.get(id = 6)
    t_suma_carg.cantidad = t_suma_carg.cantidad + mas_cargadores
    t_suma_carg.save()

    f_suma_carg = entrada_salida_acce(id_item = 6, cantidad = mas_cargadores,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_carg.save()



    return redirect('/ordenesDevueltas/')
@login_required
def anularorden(request, id):

    if not request.user.is_superuser:
        return redirect('/ordenesProcesadas/')

# Buscar id_salida 

    salida = salidasDetalle.objects.get(id_orden = id)
    id_salida = salida.id
    mas_cobras = salida.cobras
    mas_handsfree = salida.handsfree
    mas_escolta = salida.estaciones
    mas_cascos = salida.cascos
    mas_baterias = salida.baterias
    mas_cargadores = salida.cargadores

    total_radios = movimientoRadios.objects.filter(id_salida = id_salida, estado = "F").count()

    fecha_mov = datetime.date.today()

    # COBRAS 
    t_suma_cobras = inv_accesorios.objects.get(id = 1)
    t_suma_cobras.cantidad = t_suma_cobras.cantidad + mas_cobras
    t_suma_cobras.save()

    f_suma_cobras = entrada_salida_acce(id_item = 1, cantidad = mas_cobras,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_cobras.save()

    # HANDSFREE 

    t_suma_hands = inv_accesorios.objects.get (id = 2)
    t_suma_hands.cantidad = t_suma_hands.cantidad + mas_handsfree
    t_suma_hands.save()

    f_suma_hands = entrada_salida_acce(id_item = 2, cantidad = mas_handsfree,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_hands.save()

    # ESCOLTA 

    t_suma_escolta = inv_accesorios.objects.get(id = 3)
    t_suma_escolta.cantidad = t_suma_escolta.cantidad + mas_escolta
    t_suma_escolta.save()

    f_suma_escolta = entrada_salida_acce(id_item = 3, cantidad = mas_escolta,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_escolta.save() 

    # CASCOS

    t_suma_cascos = inv_accesorios.objects.get(id = 4)
    t_suma_cascos.cantidad = t_suma_cascos.cantidad + mas_cascos
    t_suma_cascos.save()

    f_suma_cascos = entrada_salida_acce(id_item = 4, cantidad = mas_cascos,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_cascos.save()

    # BATERIAS

    t_suma_bat = inv_accesorios.objects.get(id = 5)
    t_mas_baterias = mas_baterias + total_radios
    t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias
    t_suma_bat.save()

    f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_bat.save()

    # CARGADORES 

    t_suma_carg = inv_accesorios.objects.get(id = 6)
    t_suma_carg.cantidad = t_suma_carg.cantidad + mas_cargadores
    t_suma_carg.save()

    f_suma_carg = entrada_salida_acce(id_item = 6, cantidad = mas_cargadores,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
    f_suma_carg.save()

#cambiar el estatus en el inventario de radios 

    radios_inv = movimientoRadios.objects.filter(id_salida = id_salida, estado = "F")
    for i in radios_inv:
        serial = i.serial
        cambiaestado =  invSeriales.objects.get(codigo = serial)
        cambiaestado.estado_id = 4
        cambiaestado.save()

#buscar en movimientosradios todas las radios cargadas y crear un registro nueno con D

    radiosCargadas = movimientoRadios.objects.filter(id_salida = id_salida, estado = "F")
    for i in radiosCargadas:
        salida = i.id_salida
        serial = i.serial
        estado = 'D'
        id_tipo = i.id_tipo_id

        graba_radio = movimientoRadios(id_salida = salida, serial = serial, estado = estado, id_tipo_id = id_tipo)
        graba_radio.save()

# cambiar el estado de la orden 

    cambiaestadoorden = ordenRegistro.objects.filter(id = id).update(estado_id = 1)

    return redirect('/ordenesProcesadas/')
#verifica esto en produccion con el datetime 
@login_required
def salidas(request):
    fecha_actual = datetime.date.today()
    fecha_tope = datetime.date.today() + timedelta(days=2) 
    ordenes = ordenRegistro.objects.filter(estado_id = 2).order_by('fecha_entrega')
    return render(request, 'salidas.html', {"listaOrdenes": ordenes, "fecha_actal":fecha_actual, "fecha_tope":fecha_tope})
@login_required
def ordenesProcesadas(request):
    user_actual = request.user.username
    item = User.objects.filter(Q(username='fmeneses') | Q(username='jgonzalez')| Q(username='lnoriega')).order_by('id')
    if user_actual in ('ljramirez','lcastillo','jramirez','lmeneses'):
        ordenes = vista_ordenes_procesadas.objects.all().order_by('fecha_entrega')
    else:
        ordenes = vista_ordenes_procesadas.objects.filter(asignado_a = user_actual).order_by('fecha_entrega')
        # ordenes = vista_ordenes_procesadas.objects.all().order_by('fecha_entrega')
    return render(request, 'ordenesProcesadas.html', {"listaOrdenes": ordenes, "users":item})
@login_required
def ordenesCerradas(request): 
    user_actual = request.user.username
    item = User.objects.filter(Q(username='fmeneses') | Q(username='jgonzalez')| Q(username='lnoriega')).order_by('id')
    if user_actual in ('ljramirez','lcastillo','jramirez','lmeneses'):
        ordenes = vista_ordenes_cerradas.objects.all().order_by('fecha_entrega')
    else:
        ordenes = vista_ordenes_cerradas.objects.filter(asignado_a = user_actual).order_by('fecha_entrega')
        # ordenes = vista_ordenes_procesadas.objects.all().order_by('fecha_entrega')
    return render(request, 'ordenesCerradas.html', {"listaOrdenes": ordenes, "users":item})
@login_required
def ordenesDevueltas(request):

    nombre_cliente = cliente.objects.all().order_by('nombre')
    
    devueltas = vista_entrada_detalle.objects.all().order_by('-fecha_evento')
    return render(request, 'ordenesDevueltas.html', {"listaOrdenes": devueltas, "lista_cliente":nombre_cliente})
@login_required
def ordenDetalle(request, id):
    detalle = ordenRegistro.objects.get(id=id)
    color = ""
    if mochila.objects.filter(numero_orden=id).exists():
        mochila_guardada = mochila.objects.get(numero_orden=id)
        color = mochila_guardada.color
    return render(request, 'editarOrden.html', {"listaDetalles": detalle, "mochilaguardada":color})
@login_required
def detalleOrdenCerrada(request, id):
    detalle = ordenRegistro.objects.get(id=id)
    return render(request, 'editarOrden.html', {"listaDetalles": detalle})
@login_required
def generarSalida(request, id):
    
    try:
        detalle_salida = ordenRegistro.objects.get(id=id)
        msalida = salidasDetalle.objects.get(id_orden=id)
        salida_id = msalida.id
        if movimientoRadios.objects.filter(id_salida = salida_id).exists():
            cuenta_radios = movimientoRadios.objects.filter(id_salida = salida_id).count()
            serial_radios = movimientoRadios.objects.filter(id_salida = salida_id)
            form = radiotipos()
            context = {'formRadios': form}
            return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida, "datosOrden": detalle_salida, 'formRadios': form, "cantidad_radios":cuenta_radios, "serial_radios":serial_radios})
        else:
            form = radiotipos()
            context = {'formRadios': form}
            return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida, "datosOrden": detalle_salida, 'formRadios': form})
    except:
        gsalidas = salidasDetalle(id_orden=id, cobras=detalle_salida.cantidad_cobras, cargadores=detalle_salida.cantidad_cargadores,
                              handsfree=detalle_salida.cantidad_manos_libres, cascos=detalle_salida.cantidad_cascos, repetidoras=detalle_salida.cantidad_repetidoras,
                              estaciones=detalle_salida.cantidad_estaciones, baterias = detalle_salida.cantidad_baterias)
        gsalidas.save()
#####AUDITORIA############
        # ultimo_id = salidasDetalle.objects.latest('id').id
        # numero_orden = id
        # #capturar usuario actual
        # user_nombre = request.user.username
        # #fecha y hora actual
        # fecha_hora_peru = timezone.localtime(timezone.now())
        # fecha_hora_formateada = fecha_hora_peru.strftime('%Y-%m-%d %H:%M:%S')

        # log = auditoria(fecha = fecha_hora_formateada, accion = f'Genera Salida N° {ultimo_id} para la orden {numero_orden}', usuario = user_nombre)
        # log.save()

        # detalle_salida = ordenRegistro.objects.get(id=id)
        # b = detalle_salida
        # b.estado_id = 5
        # b.save() 
        form = radiotipos()
        context = {'formRadios': form}
        msalida = salidasDetalle.objects.get(id_orden=id)
        return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida, "datosOrden": detalle_salida, 'formRadios': form})
@login_required
def guardarDetalleRadio(request, id, id_orden):

    form = radiotipos()
    context = {'form':form}

    detalle_salida = ordenRegistro.objects.get(id = id_orden)
    cuenta_radios = movimientoRadios.objects.filter(id_salida = id).count()
    msalida = salidasDetalle.objects.get(id_orden = id_orden)
    serial_radios = movimientoRadios.objects.filter(id_salida = id)    
        
    if request.method == 'POST':
            
            form = radiotipos(request.POST)
            if form.is_valid():
##valida que exista 
                a = form.cleaned_data['serial']
                valida_serial = invSeriales.objects.filter(codigo = a).count()

                if valida_serial == 0:
                    tipo = form.cleaned_data['id_tipo']
                    initial_data = {
                    'serial': '',
                    'id_tipo': tipo
                    }
                    form = radiotipos(initial=initial_data)
                    return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida,"formRadios": form, "datosOrden": detalle_salida, "cantidad_radios":cuenta_radios, "serial_radios":serial_radios, 
                'error':'El serial que intenta agregar a la salida, no existe'})

##valida si la cantidad ingresada corresponde a la orden
##aqui debe llevar a otra vista que sea la salida
                if cuenta_radios >= detalle_salida.cantidad_radios: 
                     return render(request,'salidaDefinitiva.html', {"salidaGenerada": msalida, "datosOrden": detalle_salida, "cantidad_radios":cuenta_radios,"serial_radios":serial_radios})

##validar que ya no exista en la salida
                b = form.cleaned_data['serial']
                valida_serial_salida = movimientoRadios.objects.filter(serial = b, id_salida = id).count()

                if  valida_serial_salida == 1:
                    tipo = form.cleaned_data['id_tipo']
                    initial_data = {
                    'serial': '',
                    'id_tipo': tipo
                    }
                    form = radiotipos(initial=initial_data)
                    return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida,"formRadios": form, "datosOrden": detalle_salida, "cantidad_radios":cuenta_radios, "serial_radios":serial_radios,
                    'error':'Este serial ya existe en la orden'})

##validar el estado del radio
                #d = form.cleaned_data['serial']
                valida_estado = invSeriales.objects.get(codigo = a)

                if valida_estado.estado_id != 4:
                    tipo = form.cleaned_data['id_tipo']
                    initial_data = {
                    'serial': '',
                    'id_tipo': tipo
                    }
                    form = radiotipos(initial=initial_data)
                    return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida,"formRadios": form, "datosOrden": detalle_salida, "cantidad_radios":cuenta_radios, "serial_radios":serial_radios,
                'error':'El serial que intenta agregar no esta disponible'})

                else:

                    a = form.save(commit=False)
                    a.id_salida = id
                    a.estado = 'F'
                    a.save()

##cambia el estado del radio despues de guardarlo para que no este disponible
                    c = form.cleaned_data['serial']
                    cambiaestado =  invSeriales.objects.get(codigo = c)
                    cambiaestado.estado_id = 5
                    cambiaestado.save()

                    detalle_salida = ordenRegistro.objects.get(id=id_orden)        
                    cuenta_radios = movimientoRadios.objects.filter(id_salida = id).count()
                    serial_radios = movimientoRadios.objects.filter(id_salida = id)
                    msalida = salidasDetalle.objects.get(id_orden=id_orden)

#aqui deberia generar el ticket e imprimir otra vista
# Comprueba si ya se ha llegado a la cantidad de radios de la orden  
#                   # cambia el estado de la orden automaticamente al completar la cantidad de radios
                    if cuenta_radios == detalle_salida.cantidad_radios:
                        ##cierra la orden
                        detalle_salida = ordenRegistro.objects.get(id=id_orden)
                        e = detalle_salida
                        e.estado_id = 5
                        e.save()
                        ############ AQUI DEBE DESCONTAR DE LOS ACCESORIOS #################
                        # TIENE QUE TOMAR EN CUENTA LAS BATERIAS QUE SE VAN PUESTAS EN LOS RADIOS + ADICIONALES #######
                        #
                        menos_cobras = e.cantidad_cobras
                        menos_handsfree = e.cantidad_manos_libres
                        menos_escolta = e.cantidad_estaciones
                        menos_cascos = e.cantidad_cascos
                        menos_baterias = e.cantidad_baterias
                        menos_cargadores = e.cantidad_cargadores

                        fecha_mov = datetime.date.today()

                        # COBRAS 
                        t_descuento_cobras = inv_accesorios.objects.get(id = 1)
                        t_descuento_cobras.cantidad = t_descuento_cobras.cantidad - menos_cobras
                        t_descuento_cobras.save()

                        f_descuento_cobras = entrada_salida_acce(id_item = 1, cantidad = menos_cobras,
                                                                  tipo_mov = 'S', fecha_mov = fecha_mov)
                        f_descuento_cobras.save()

                        # HANDSFREE 

                        t_descuento_hands = inv_accesorios.objects.get (id = 2)
                        t_descuento_hands.cantidad = t_descuento_hands.cantidad - menos_handsfree
                        t_descuento_hands.save()

                        f_descuento_hands = entrada_salida_acce(id_item = 2, cantidad = menos_handsfree,
                                                                  tipo_mov = 'S', fecha_mov = fecha_mov)
                        f_descuento_hands.save()

                        # ESCOLTA 

                        t_descuento_escolta = inv_accesorios.objects.get(id = 3)
                        t_descuento_escolta.cantidad = t_descuento_escolta.cantidad - menos_escolta
                        t_descuento_escolta.save()

                        f_descuento_escolta = entrada_salida_acce(id_item = 3, cantidad = menos_escolta,
                                                                  tipo_mov = 'S', fecha_mov = fecha_mov)
                        f_descuento_escolta.save()
 

                        # CASCOS

                        t_descuento_cascos = inv_accesorios.objects.get(id = 4)
                        t_descuento_cascos.cantidad = t_descuento_cascos.cantidad - menos_cascos
                        t_descuento_cascos.save()

                        f_descuento_cascos = entrada_salida_acce(id_item = 4, cantidad = menos_cascos,
                                                                  tipo_mov = 'S', fecha_mov = fecha_mov)
                        f_descuento_cascos.save()

                        # BATERIAS

                        t_descuento_bat = inv_accesorios.objects.get(id = 5)
                        t_baterias = menos_baterias + cuenta_radios
                        t_descuento_bat.cantidad = t_descuento_bat.cantidad - t_baterias
                        t_descuento_bat.save()

                        f_descuento_bat = entrada_salida_acce(id_item = 5, cantidad = t_baterias,
                                                                  tipo_mov = 'S', fecha_mov = fecha_mov)
                        f_descuento_bat.save()

                        # CARGADORES 

                        t_descuento_carg = inv_accesorios.objects.get(id = 6)
                        t_descuento_carg.cantidad = t_descuento_carg.cantidad - menos_cargadores
                        t_descuento_carg.save()

                        f_descuento_carg = entrada_salida_acce(id_item = 6, cantidad = menos_cargadores,
                                                                  tipo_mov = 'S', fecha_mov = fecha_mov)
                        f_descuento_carg.save() 
                        return redirect('ordenesProcesadas')
                    
                    tipo = form.cleaned_data['id_tipo']
                    initial_data = {
                    'serial': '',
                    'id_tipo': tipo
                    }
                    form = radiotipos(initial=initial_data)
                    return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida, "datosOrden": detalle_salida, "formRadios": form, "cantidad_radios":cuenta_radios,"serial_radios":serial_radios})
                    
    else:
        # form = radiotipos(initial={'serial': ''})
        return render(request, 'salidaGenerada.html', {"salidaGenerada": msalida, "datosOrden": detalle_salida, "formRadios": form, "cantidad_radios":cuenta_radios,"serial_radios":serial_radios})
@login_required           
def consultainventario(request):

    inv_disp = invSeriales.objects.filter(estado_id =4)
    cuenta_disponibles = invSeriales.objects.filter(estado_id =4).count()
    inv_nodisp = invSeriales.objects.filter(estado_id =5)
    cuenta_nodisponibles = invSeriales.objects.filter(estado_id =5).count()
    inv_malogrado = invSeriales.objects.filter(estado_id =6)
    cuenta_malogrado = invSeriales.objects.filter(estado_id =6).count()
    amarillas_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 1).count()
    moradas_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 2).count()
    plomo_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 3).count()
    amarillas_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 1).count()
    moradas_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 2).count()
    plomo_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 3).count()
    extraviadas = invSeriales.objects.filter(estado_id =7).count()





    form = agregarInven()
    context = {'form':form}

    if request.method == 'POST':
        form = agregarInven(request.POST)
        
        if form.is_valid():
            b = form.cleaned_data['codigo']
            if invSeriales.objects.filter(codigo =b).exists():
                return render(request, 'inventario.html', {"form": form,"listaDisponibles":inv_disp, "listaNoDisponible":inv_nodisp, "listaMalogrado":inv_malogrado, "cantidadDisponible":cuenta_disponibles,
                "cantidadnoDispo":cuenta_nodisponibles,"cantidadMalogrado":cuenta_malogrado, "buscaradios":formBuscaRadio, "amarillasDisponibles":amarillas_disp, 
                "moradasDisponibles":moradas_disp, "plomoDisponibles":plomo_disp, "amarillasnoDisponibles":amarillas_nodisp, "moradasnoDisponibles":moradas_nodisp, 
                "plomonoDisponibles":plomo_nodisp})
            else:           
                form.save() 
                return render(request,'inventario.html', {"form": form,"listaDisponibles":inv_disp, "listaNoDisponible":inv_nodisp, "listaMalogrado":inv_malogrado, "cantidadDisponible":cuenta_disponibles,
                "cantidadnoDispo":cuenta_nodisponibles, "cantidadMalogrado":cuenta_malogrado, "buscaradios":formBuscaRadio, "amarillasDisponibles":amarillas_disp, 
                "moradasDisponibles":moradas_disp, "plomoDisponibles":plomo_disp, "amarillasnoDisponibles":amarillas_nodisp, "moradasnoDisponibles":moradas_nodisp, 
                "plomonoDisponibles":plomo_nodisp, 'extraviadas':extraviadas})         

    return render(request,'inventario.html', {"form": form,"listaDisponibles":inv_disp, "listaNoDisponible":inv_nodisp, "listaMalogrado":inv_malogrado, "cantidadDisponible":cuenta_disponibles,
    "cantidadnoDispo":cuenta_nodisponibles,"cantidadMalogrado":cuenta_malogrado,"buscaradios":formBuscaRadio, "amarillasDisponibles":amarillas_disp, 
    "moradasDisponibles":moradas_disp, "plomoDisponibles":plomo_disp, "amarillasnoDisponibles":amarillas_nodisp, "moradasnoDisponibles":moradas_nodisp, 
                "plomonoDisponibles":plomo_nodisp, 'extraviadas':extraviadas})

@login_required
def movimientosRadios(request):


    inv_disp = invSeriales.objects.filter(estado_id =4)
    cuenta_disponibles = invSeriales.objects.filter(estado_id =4).count()
    inv_nodisp = invSeriales.objects.filter(estado_id =5)
    cuenta_nodisponibles = invSeriales.objects.filter(estado_id =5).count()
    inv_malogrado = invSeriales.objects.filter(estado_id =6)
    cuenta_malogrado = invSeriales.objects.filter(estado_id =6).count()
    amarillas_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 1).count()
    moradas_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 2).count()
    plomo_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 3).count()
    amarillas_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 1).count()
    moradas_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 2).count()
    plomo_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 3).count()
    extraviadas = invSeriales.objects.filter(estado_id =7).count()

    form = agregarInven()
    context = {'form':form}

    formBuscarx = formBuscaRadio()

    if request.method == 'POST':

        serial = request.POST['serial']
        valida_serial = invSeriales.objects.filter(codigo = serial).count()
        if valida_serial == 0:
            return render(request,'inventario.html', {"form": form,"listaDisponibles":inv_disp, "listaNoDisponible":inv_nodisp, "listaMalogrado":inv_malogrado, "cantidadDisponible":cuenta_disponibles,
                "cantidadnoDispo":cuenta_nodisponibles, "cantidadMalogrado":cuenta_malogrado, "buscaradios":formBuscaRadio, "amarillasDisponibles":amarillas_disp, 
                "moradasDisponibles":moradas_disp, "plomoDisponibles":plomo_disp, "amarillasnoDisponibles":amarillas_nodisp, "moradasnoDisponibles":moradas_nodisp, 
                "plomonoDisponibles":plomo_nodisp, 'extraviadas':extraviadas})         
        else:
            estadorx = invSeriales.objects.get(codigo = serial)
            ultima_salida = movimientoRadios.objects.filter(serial = serial)
            ultima_fecha = ultima_salida.aggregate(fecha_maxima=models.Max('fecha_creacion'))['fecha_maxima']
            if ultima_fecha is not None:
                formatedDate = ultima_fecha.strftime("%d/%m/%Y %H:%M:%S")

                ultimo_cliente = movimientoRadios.objects.get(fecha_creacion = ultima_fecha, serial = serial)
                salida_id = ultimo_cliente.id_salida

                detalle_id = salidasDetalle.objects.get(id = salida_id)
                
                numero_orden = detalle_id.id_orden
                #verifica aqui
                orden_cliente = ordenRegistro.objects.get(id = numero_orden)
                nombre_cliente = orden_cliente.cliente
                
            else:
                formatedDate = 'aun no tiene salidas'
                nombre_cliente = 'aun no tiene salidas'

        return render(request,'inventario.html', {"form": form,"listaDisponibles":inv_disp, "listaNoDisponible":inv_nodisp, "listaMalogrado":inv_malogrado, "cantidadDisponible":cuenta_disponibles,
                "cantidadnoDispo":cuenta_nodisponibles, "cantidadMalogrado":cuenta_malogrado, "buscaradios":formBuscaRadio, "amarillasDisponibles":amarillas_disp, 
                "moradasDisponibles":moradas_disp, "plomoDisponibles":plomo_disp, "amarillasnoDisponibles":amarillas_nodisp, "moradasnoDisponibles":moradas_nodisp, 
                "plomonoDisponibles":plomo_nodisp, 'extraviadas':extraviadas, "estado":estadorx.estado,"ultima_salida":formatedDate, "cliente":nombre_cliente})         
@login_required
def cambiaEstadoEntregado(request, id):
    detalle_salida = ordenRegistro.objects.get(id=id)
    b = detalle_salida
    b.estado_id = 3
    b.save() 

    #####AUDITORIA############
    # numero_orden = id
    # #capturar usuario actual
    # user_nombre = request.user.username
    # #fecha y hora actual
    # fecha_hora_peru = timezone.localtime(timezone.now())
    # fecha_hora_formateada = fecha_hora_peru.strftime('%Y-%m-%d %H:%M:%S')

    # log = auditoria(fecha = fecha_hora_formateada, accion = f'Cambia estado a ENTREGADO a la orden N° {numero_orden}', usuario = user_nombre)
    # log.save()



    # ordenes = ordenRegistro.objects.filter(estado_id = 3)
    ordenes = vista_ordenes_procesadas.objects.all().order_by('fecha_entrega')
    return render(request, 'ordenesProcesadas.html', {"listaOrdenes": ordenes})
@login_required
def generarPDFPreparados(request, id):

    ordenes = ordenRegistro.objects.get(id = id)
    b = ordenes

    cliente = str(b.cliente)
    radios = str(b.cantidad_radios)
    cobras = str(b.cantidad_cobras)
    baterias = str(b.cantidad_baterias)
    cargadores = str(b.cantidad_cargadores)
    manos_libres = str(b.cantidad_manos_libres)
    cascos = str(b.cantidad_cascos)
    repetidoras = str(b.cantidad_repetidoras)
    estaciones = str(b.cantidad_estaciones)
    fentrega = str(b.fecha_entrega)
    pedido = str(b.id)

    buffer = BytesIO()
    tamano_pagina =  (20*cm, 10*cm)
    pdf = canvas.Canvas(buffer, pagesize=tamano_pagina)
   

    pdf.drawString(2*cm, 9*cm,"N° Pedido: "+ pedido)
    pdf.drawString(2*cm, 8*cm, "Cliente: "+ cliente)
    pdf.drawString(2*cm, 7*cm,"Fecha entrega: " +fentrega)
    if b.cantidad_radios > 0:
        pdf.drawString(2*cm, 6*cm, "Radios: "+ radios)
    if b.cantidad_cobras > 0:
        pdf.drawString(2*cm, 5.5*cm, "Cobras: "+ cobras)
    if b.cantidad_baterias > 0:
        pdf.drawString(2*cm, 5*cm, "Baterias extra: "+ baterias)
    if b.cantidad_cargadores > 0:
        pdf.drawString(2*cm, 4.5*cm, "Cargadores: " + cargadores)
    if b.cantidad_manos_libres > 0:
        pdf.drawString(2*cm, 4*cm,"Manos libres: " + manos_libres)
    if b.cantidad_cascos > 0:
        pdf.drawString(2*cm, 3.5*cm,"Cascos: " + cascos)
    if b.cantidad_repetidoras > 0:
        pdf.drawString(2*cm, 3*cm,"Repetidoras: " + repetidoras)
    if b.cantidad_estaciones > 0:
        pdf.drawString(2*cm, 2.5*cm,"Estaciones: " + estaciones)
    
    
    pdf.showPage()

    pdf.save()

    buffer.seek(0)
    nombre_archivo = str(id)+'_'+str(cliente)+'.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename= {nombre_archivo}'


    return response
@login_required
def generarEntrada(request, id, orden_id):

    if request.method == 'POST':
        formEntrada = formEntradaDetalle(request.POST)

        if formEntrada.is_valid():

            if accesoriosFaltantes.objects.filter(id_salida = id).exists():
                cobras = formEntrada.cleaned_data['cobras']
                baterias = formEntrada.cleaned_data['baterias']
                cargadores = formEntrada.cleaned_data['cargadores']
                handsfree = formEntrada.cleaned_data['handsfree']
                cascos = formEntrada.cleaned_data['cascos']
                repetidoras = formEntrada.cleaned_data['repetidoras']
                estaciones = formEntrada.cleaned_data['estaciones']
                observaciones = formEntrada.cleaned_data['observaciones']

                actualiza_acce = accesoriosFaltantes.objects.get(id_salida = id)
                actualiza_acce.fcobras = actualiza_acce.fcobras - cobras
                actualiza_acce.fbaterias = actualiza_acce.fbaterias - baterias
                actualiza_acce.fcargadores = actualiza_acce.fcargadores - cargadores
                actualiza_acce.fhandsfree = actualiza_acce.fhandsfree - handsfree
                actualiza_acce.fcascos = actualiza_acce.fcascos - cascos
                actualiza_acce.frepetidoras = actualiza_acce.frepetidoras - repetidoras
                actualiza_acce.festaciones = actualiza_acce.festaciones - estaciones
                actualiza_acce.save()

                ### ENTRADA ACCESORIOS ###

                mas_cobras = cobras
                mas_handsfree = handsfree
                mas_escolta = estaciones
                mas_cascos = cascos
                mas_baterias = baterias
                mas_cargadores = cargadores

                fecha_mov = datetime.date.today()

                # COBRAS 
                t_suma_cobras = inv_accesorios.objects.get(id = 1)
                t_suma_cobras.cantidad = t_suma_cobras.cantidad + mas_cobras
                t_suma_cobras.save()

                f_suma_cobras = entrada_salida_acce(id_item = 1, cantidad = mas_cobras,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_cobras.save()

                # HANDSFREE 

                t_suma_hands = inv_accesorios.objects.get (id = 2)
                t_suma_hands.cantidad = t_suma_hands.cantidad + mas_handsfree
                t_suma_hands.save()

                f_suma_hands = entrada_salida_acce(id_item = 2, cantidad = mas_handsfree,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_hands.save()

                # ESCOLTA 

                t_suma_escolta = inv_accesorios.objects.get(id = 3)
                t_suma_escolta.cantidad = t_suma_escolta.cantidad + mas_escolta
                t_suma_escolta.save()

                f_suma_escolta = entrada_salida_acce(id_item = 3, cantidad = mas_escolta,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_escolta.save() 

                # CASCOS

                t_suma_cascos = inv_accesorios.objects.get(id = 4)
                t_suma_cascos.cantidad = t_suma_cascos.cantidad + mas_cascos
                t_suma_cascos.save()

                f_suma_cascos = entrada_salida_acce(id_item = 4, cantidad = mas_cascos,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_cascos.save()

                # BATERIAS

                t_suma_bat = inv_accesorios.objects.get(id = 5)
                t_mas_baterias = mas_baterias 
                t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias 
                t_suma_bat.save()

                f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_bat.save()

                # CARGADORES 

                t_suma_carg = inv_accesorios.objects.get(id = 6)
                t_suma_carg.cantidad = t_suma_carg.cantidad + mas_cargadores
                t_suma_carg.save()

                f_suma_carg = entrada_salida_acce(id_item = 6, cantidad = mas_cargadores,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_carg.save()


                actualiza_acce = accesoriosFaltantes.objects.get(id_salida = id) 
                if actualiza_acce.fcobras + actualiza_acce.fbaterias + actualiza_acce.fcargadores + actualiza_acce.fhandsfree + actualiza_acce.fcascos + actualiza_acce.frepetidoras + actualiza_acce.festaciones == 0:
                    actualiza_acce.delete()
                    return redirect('/verFaltante/')
                return redirect('/verFaltante/')

            #tomo cantidades enviadas desde form
            cobras = formEntrada.cleaned_data['cobras']
            baterias = formEntrada.cleaned_data['baterias']
            cargadores = formEntrada.cleaned_data['cargadores']
            handsfree = formEntrada.cleaned_data['handsfree']
            cascos = formEntrada.cleaned_data['cascos']
            repetidoras = formEntrada.cleaned_data['repetidoras']
            estaciones = formEntrada.cleaned_data['estaciones']
            observaciones = formEntrada.cleaned_data['observaciones']

            #busca las cantidades en la salida para comparar
            datosSalida = salidasDetalle.objects.get(id = id)
            numeroOrden = datosSalida.id_orden
            cobrasSal = datosSalida.cobras
            bateriasSal = datosSalida.baterias
            cargadoresSal = datosSalida.cargadores
            handsfreeSal = datosSalida.handsfree
            cascosSal = datosSalida.cascos
            repetidorasSal = datosSalida.repetidoras
            estacionesSal = datosSalida.estaciones

            ordenes = ordenRegistro.objects.get(id = orden_id)
            msalida = salidasDetalle.objects.get(id_orden=orden_id)
            radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")
            cuentaRxcargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D").count()
            cuentaRxOrden =  movimientoRadios.objects.filter(id_salida = id, estado = "F").count()
            formEntrada = formEntradaDetalle()
            form = guardaEntradaRx()

            #validar las cantidades no deben ser mayores 

            if cobras > cobrasSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})            
            if baterias > bateriasSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})
            if cargadores > cargadoresSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})
            if handsfree > handsfreeSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})
            if cascos > cascosSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})
            if repetidoras > repetidorasSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})
            if estaciones > estacionesSal :
                return render(request, 'entradas.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, "error1":"La cantidades ingresadas no pueden ser mayores a las registradas en el pedido. Verifique"})
            
            # crear el registro de lo faltante (accesorios)

            faltantecobras = cobrasSal - cobras
            faltantebat = bateriasSal - baterias
            faltantecarga = cargadoresSal - cargadores
            faltantehands = handsfreeSal - handsfree
            faltantecascos = cascosSal - cascos
            faltanterep = repetidorasSal - repetidoras
            faltanteest = estacionesSal - estaciones

            sumafaltante = faltantecobras + faltantebat + faltantecarga + faltantehands + faltantecascos + faltanterep + faltanteest

            if sumafaltante > 0:
                guardafaltante = accesoriosFaltantes(id_salida = id, fcobras = faltantecobras, fbaterias = faltantebat, fcargadores = faltantecarga,
                                                     fhandsfree = faltantehands, fcascos = faltantecascos, frepetidoras = faltanterep,
                                                     festaciones = faltanteest)
                
                guardafaltante.save()

                mas_cobras = cobras
                mas_handsfree = handsfree
                mas_escolta = estaciones
                mas_cascos = cascos
                mas_baterias = baterias
                mas_cargadores = cargadores

                fecha_mov = datetime.date.today()

                # COBRAS 
                t_suma_cobras = inv_accesorios.objects.get(id = 1)
                t_suma_cobras.cantidad = t_suma_cobras.cantidad + mas_cobras
                t_suma_cobras.save()

                f_suma_cobras = entrada_salida_acce(id_item = 1, cantidad = mas_cobras,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_cobras.save()

                # HANDSFREE 

                t_suma_hands = inv_accesorios.objects.get (id = 2)
                t_suma_hands.cantidad = t_suma_hands.cantidad + mas_handsfree
                t_suma_hands.save()

                f_suma_hands = entrada_salida_acce(id_item = 2, cantidad = mas_handsfree,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_hands.save()

                # ESCOLTA 

                t_suma_escolta = inv_accesorios.objects.get(id = 3)
                t_suma_escolta.cantidad = t_suma_escolta.cantidad + mas_escolta
                t_suma_escolta.save()

                f_suma_escolta = entrada_salida_acce(id_item = 3, cantidad = mas_escolta,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_escolta.save() 

                # CASCOS

                t_suma_cascos = inv_accesorios.objects.get(id = 4)
                t_suma_cascos.cantidad = t_suma_cascos.cantidad + mas_cascos
                t_suma_cascos.save()

                f_suma_cascos = entrada_salida_acce(id_item = 4, cantidad = mas_cascos,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_cascos.save()

                # BATERIAS

                t_suma_bat = inv_accesorios.objects.get(id = 5)
                t_mas_baterias = mas_baterias 
                t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias 
                t_suma_bat.save()

                f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_bat.save()

                # CARGADORES 

                t_suma_carg = inv_accesorios.objects.get(id = 6)
                t_suma_carg.cantidad = t_suma_carg.cantidad + mas_cargadores
                t_suma_carg.save()

                f_suma_carg = entrada_salida_acce(id_item = 6, cantidad = mas_cargadores,
                                                                    tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_carg.save()


            # crear el registro de lo faltante (radios)

            rxOrden = movimientoRadios.objects.filter(id_salida = id, estado = "F")

            for i in rxOrden:
                serfaltante = i.serial
                bat_faltante_rx = 0

                if not movimientoRadios.objects.filter(serial = serfaltante, id_salida = id, estado = "D").exists():
                    guardafaltanterx = radiosFantantes()
                    guardafaltanterx.id_salida = id
                    guardafaltanterx.fserial = serfaltante
                    bat_faltante_rx += 1
                    guardafaltanterx.save()

            numorden = salidasDetalle.objects.get(id_orden=orden_id)
            numordena = numorden.id_orden
            client = ordenRegistro.objects.get(id=numordena)
            client_id = client.cliente
            client.estado_id = 4
            client.save()

            if not entradaDetalle.objects.filter(id_salida = id).exists():
                guardaEntrada = entradaDetalle(id_salida = id, id_orden = numordena, cobras = cobras, baterias = baterias, cargadores = cargadores,
                                           handsfree = handsfree, cascos = cascos, repetidoras = repetidoras, estaciones = estaciones, cliente = client_id,
                                            observaciones = observaciones)
            
                guardaEntrada.save()

            ### ENTRADA ACCESORIOS ###

            mas_cobras = cobras
            mas_handsfree = handsfree
            mas_escolta = estaciones
            mas_cascos = cascos
            mas_baterias = baterias
            mas_cargadores = cargadores

            fecha_mov = datetime.date.today()

            # COBRAS 
            t_suma_cobras = inv_accesorios.objects.get(id = 1)
            t_suma_cobras.cantidad = t_suma_cobras.cantidad + mas_cobras
            t_suma_cobras.save()

            f_suma_cobras = entrada_salida_acce(id_item = 1, cantidad = mas_cobras,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
            f_suma_cobras.save()

            # HANDSFREE 

            t_suma_hands = inv_accesorios.objects.get (id = 2)
            t_suma_hands.cantidad = t_suma_hands.cantidad + mas_handsfree
            t_suma_hands.save()

            f_suma_hands = entrada_salida_acce(id_item = 2, cantidad = mas_handsfree,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
            f_suma_hands.save()

            # ESCOLTA 

            t_suma_escolta = inv_accesorios.objects.get(id = 3)
            t_suma_escolta.cantidad = t_suma_escolta.cantidad + mas_escolta
            t_suma_escolta.save()

            f_suma_escolta = entrada_salida_acce(id_item = 3, cantidad = mas_escolta,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
            f_suma_escolta.save() 

            # CASCOS

            t_suma_cascos = inv_accesorios.objects.get(id = 4)
            t_suma_cascos.cantidad = t_suma_cascos.cantidad + mas_cascos
            t_suma_cascos.save()

            f_suma_cascos = entrada_salida_acce(id_item = 4, cantidad = mas_cascos,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
            f_suma_cascos.save()

            # BATERIAS

            t_suma_bat = inv_accesorios.objects.get(id = 5)
            t_mas_baterias = mas_baterias - bat_faltante_rx
            t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias 
            t_suma_bat.save()

            f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
            f_suma_bat.save()

            # CARGADORES 

            t_suma_carg = inv_accesorios.objects.get(id = 6)
            t_suma_carg.cantidad = t_suma_carg.cantidad + mas_cargadores
            t_suma_carg.save()

            f_suma_carg = entrada_salida_acce(id_item = 6, cantidad = mas_cargadores,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
            f_suma_carg.save()


                #####AUDITORIA############
                # numero_orden = numordena
                # #capturar usuario actual
                # user_nombre = request.user.username
                # #fecha y hora actual
                # fecha_hora_peru = timezone.localtime(timezone.now())
                # fecha_hora_formateada = fecha_hora_peru.strftime('%Y-%m-%d %H:%M:%S')

                # log = auditoria(fecha = fecha_hora_formateada, accion = f'Pedido recibido de salida N° {id} de la orden {numero_orden}', usuario = user_nombre)
                # log.save()

            return redirect('/ordenesCerradas/')
@login_required
def verFaltante(request):
    
    rx_listado_faltante = vista_radios_faltantes.objects.all()
    listadoFantante = vista_accesorios_faltantes.objects.all()
    return render(request,"faltantes.html",{"listAccFaltante":listadoFantante, "listadorxfaltante":rx_listado_faltante})
@login_required
def generarPDFGuia(request, id):

    # font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'MS_Reference_Sans_Serif.ttf')
    # pdfmetrics.registerFont(TTFont('MS_Reference_Sans_Serif', font_path))
    # pdfmetrics.registerFontFamily('MS_Reference_Sans_Serif', normal='MS_Reference_Sans_Serif')
    
    
    detalle_acce = salidasDetalle.objects.get(id_orden = id)
    b = detalle_acce

    orden = b.id_orden
    salida_id = b.id

    ordenes = ordenRegistro.objects.get(id = orden)

    c = ordenes

    if mochila.objects.filter(numero_orden = orden).exists(): 
        color_mochila = mochila.objects.get(numero_orden = orden)
        color_descrip = color_mochila.color
    else:
        color_descrip = ""

    fecha_actual = datetime.date.today().strftime('%d/%m/%Y')

    radios = movimientoRadios.objects.filter(id_salida = salida_id, estado = 'F').values_list('serial', flat=True)

    d = radios

    rx_amarillas = vista_movimiento_radios_tipos.objects.filter(id_salida = salida_id, tipo = 'EP 450 (Amarillas)').count()
    rx_moradas =  vista_movimiento_radios_tipos.objects.filter(id_salida = salida_id, tipo = 'DEP 450 (Moradas)').count()
    rx_plomo = vista_movimiento_radios_tipos.objects.filter(id_salida = salida_id, tipo = 'DEP 450 (Plomo)').count()

    amarillas = rx_amarillas
    moradas = rx_moradas
    plomo = rx_plomo



    cliente = str(c.cliente)
    cobras = str(b.cobras)
    cargadores = str(b.cargadores)
    handsfree = str(b.handsfree)
    cascos = str(b.cascos)
    estaciones = str(b.estaciones)
    repetidoras = str(b.repetidoras)
    baterias = str(b.baterias)
    fecha_entrega = str(c.fecha_entrega)
    fecha_evento = (c.fecha_evento_desde).strftime('%d/%m/%Y')
    direccion = str(c.direccion_entrega)
    rx = d

    buffer = BytesIO()

    pdf = canvas.Canvas(buffer)
    ancho_pagina, altura_pagina = letter = (21.59*cm, 27.94*cm)
    tamano_direccion = len(direccion)
    
    if tamano_direccion > 44:
        linea_dir1 = direccion[:44]
        linea_dir2 = direccion[45:88]
        linea_dir3 = direccion[89:133]
    else:
        linea_dir1 = direccion[:44]
        linea_dir2 = ""
        linea_dir3 = "" 

    pdf.drawString(18*cm, altura_pagina - 6.05*cm, str(id))
    pdf.drawString(5*cm, altura_pagina - 9.35*cm, str(cliente))
    pdf.drawString(3.5*cm, altura_pagina - 10*cm, str(linea_dir1))
    pdf.drawString(3.5*cm, altura_pagina - 10.4*cm, str(linea_dir2))
    pdf.drawString(3.5*cm, altura_pagina - 10.8*cm, str(linea_dir3))
    pdf.drawString(15.5*cm, altura_pagina - 9.35*cm, str(fecha_actual))
    pdf.drawString(14*cm, altura_pagina - 10*cm, 'Evento: '+ str(fecha_evento))
    l = 60
    pdf.drawString(l, 50, str(color_descrip))
    if b.cobras > 0:
        pdf.drawString(14*cm, altura_pagina - 15.25*cm, str(cobras) + ' UND')
    if b.handsfree > 0:
        pdf.drawString(14*cm, altura_pagina - 15.95*cm, str(handsfree) + ' UND')
    if b.cascos > 0:
        pdf.drawString(5*cm, altura_pagina - 17.45*cm, str('Headset Antiruido'))
        pdf.drawString(14*cm, altura_pagina - 17.45*cm, str(cascos + ' UND'))
    if b.baterias > 0:
        pdf.drawString(14*cm, altura_pagina - 19.45*cm, str(baterias) + ' UND')
    if b.cargadores > 0:
        pdf.drawString(14*cm, altura_pagina - 20.15*cm, str(cargadores + ' UND'))
    if b.estaciones >0:
        pdf.drawString(14*cm, altura_pagina - 16.75*cm, str(estaciones + ' UND'))
    if b.repetidoras > 0:
        pdf.drawString(5*cm, altura_pagina - 20.85*cm, str('Repetidoras'))
        pdf.drawString(14*cm, altura_pagina - 20.85*cm, str(repetidoras + ' UND'))
    #radios
    if rx_amarillas > 0:
        pdf.drawString(14*cm, altura_pagina - 11.95*cm, str(amarillas))
        pdf.drawString(14.5*cm, altura_pagina - 11.95*cm, str(' UND'))
    if rx_moradas > 0:
        pdf.drawString(14*cm, altura_pagina - 13.45*cm, str(moradas))
        pdf.drawString(14.5*cm, altura_pagina - 13.45*cm, str(' UND'))
    if rx_plomo > 0:
        pdf.drawString(14*cm, altura_pagina - 12.65*cm, str(plomo))
        pdf.drawString(14.5*cm, altura_pagina - 12.65*cm, str(' UND'))
    # #y = 110
    font_size = 10
    pdf.setFont("Helvetica", font_size)
    x = 60
    h = 60
    f = 60
    g = 60
    # for i in rx:
    #     pdf.drawString(x, 110, str(i))
    #     #y += 20
    #     x += 30
    cuenta_serial = len(rx)
    if cuenta_serial > 16:
        for indice, i in enumerate(rx):
            if indice <=15:
                pdf.drawString(x, 90, str(i))
                x += 30
            if indice >15 and indice <31:
                pdf.drawString(h, 80, str(i))
                h += 30
            if indice >=31 and indice <47:
                pdf.drawString(f, 70, str(i))
                f += 30
            if indice >=47 and indice <60:
                pdf.drawString(g, 60, str(i))
                g += 30
    else:
        for i in rx[:cuenta_serial]:
            pdf.drawString(x, 90, str(i))
            #y += 20
            x += 30

    

    pdf.showPage()

    pdf.save()

    buffer.seek(0)

    nombre_archivo = str(id)+'_'+str(fecha_entrega)+'.pdf'
   
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename= {nombre_archivo}'


    return response
@login_required
def guardarcolormochila(request, id):

    if request.method == 'POST':

        if mochila.objects.filter(numero_orden=id).exists():
            mochila_color = request.POST.get('mochila')
            mochila_guardada = mochila.objects.get(numero_orden=id)
            mochila_guardada.color = mochila_color
            mochila_guardada.save()
        else:
            mochila_color = request.POST.get('mochila')
            guardamochila = mochila()
            guardamochila.numero_orden = id
            guardamochila.color = mochila_color 
            guardamochila.save()

        color = ""
        if mochila.objects.filter(numero_orden=id).exists():
            mochila_guardada = mochila.objects.get(numero_orden=id)
            color = mochila_guardada.color

        detalle = ordenRegistro.objects.get(id=id)
        return render(request, 'editarOrden.html', {"listaDetalles": detalle, "mochilaguardada":color})
@login_required
def pdfguiadevueltos(request, id):

    # font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'times.ttf')
    # pdfmetrics.registerFont(TTFont('TimesNewRoman', font_path))
    # pdfmetrics.registerFontFamily('TimesNewRoman', normal='TimesNewRoman')

    detalle_acce = salidasDetalle.objects.get(id_orden = id)
    b = detalle_acce

    orden = b.id_orden
    salida_id = b.id

    ordenes = ordenRegistro.objects.get(id = orden)

    c = ordenes

    if mochila.objects.filter(numero_orden = orden).exists(): 
        color_mochila = mochila.objects.get(numero_orden = orden)
        color_descrip = color_mochila.color
    else:
        color_descrip = ""

    fecha_actual = datetime.date.today().strftime('%d/%m/%Y')

    radios = movimientoRadios.objects.filter(id_salida = salida_id, estado = 'F').values_list('serial', flat=True)

    d = radios

    rx_amarillas = vista_movimiento_radios_tipos.objects.filter(id_salida = salida_id, tipo = 'EP 450 (Amarillas)', estado = 'F').count()
    rx_moradas =  vista_movimiento_radios_tipos.objects.filter(id_salida = salida_id, tipo = 'DEP 450 (Moradas)', estado = 'F').count()
    rx_plomo = vista_movimiento_radios_tipos.objects.filter(id_salida = salida_id, tipo = 'DEP 450 (Plomo)', estado = 'F').count()

    amarillas = rx_amarillas
    moradas = rx_moradas
    plomo = rx_plomo



    cliente = str(c.cliente)
    cobras = str(b.cobras)
    cargadores = str(b.cargadores)
    handsfree = str(b.handsfree)
    cascos = str(b.cascos)
    estaciones = str(b.estaciones)
    repetidoras = str(b.repetidoras)
    baterias = str(b.baterias)
    fecha_entrega = str(c.fecha_entrega)
    fecha_evento = (c.fecha_evento_desde).strftime('%d/%m/%Y')
    if c.fecha_evento_hasta is None:
        fecha_evento_hasta = ""
    else:
        fecha_evento_hasta = (c.fecha_evento_hasta).strftime('%d/%m/%Y')
    direccion = str(c.direccion_entrega)
    rx = d

    buffer = BytesIO()

    pdf = canvas.Canvas(buffer)
    ancho_pagina, altura_pagina = letter = (21.59*cm, 27.94*cm)
    tamano_direccion = len(direccion)

    image_filename = 'logo02.jpg'
    image1_filename = 'logo_22.jpg'
    image_path = f'radios/static/{image_filename}'
    image_path1 = f'radios/static/{image1_filename}'


    # Calcula el tamaño y posición de la imagen
    image_width = 60
    image_height = 60
    o = 60
    z = letter[1] - image_height

    image_width01 = 300
    image_height01 = 100

    o1 = 60
    z2 = z - image_height -20
    

    # Agrega la imagen al documento PDF
    #pdf.drawImage(image_path, o, z, width=image_width, height=image_height)
    pdf.drawImage(image_path1, o1, z2, width=image_width01, height=image_height01)
    
    if tamano_direccion > 44:
        linea_dir1 = direccion[:44]
        linea_dir2 = direccion[45:88]
        linea_dir3 = direccion[89:133]
    else:
        linea_dir1 = direccion[:44]
        linea_dir2 = ""
        linea_dir3 = "" 

    pdf.drawString(18*cm, altura_pagina - 6.05*cm, 'Salida N°: '+ str(id))
    pdf.drawString(3.5*cm, altura_pagina - 9.35*cm, 'Cliente: '+ str(cliente))
    pdf.drawString(3.5*cm, altura_pagina - 10*cm, 'Direccion: ' + str(linea_dir1))
    pdf.drawString(3.5*cm, altura_pagina - 10.4*cm, str(linea_dir2))
    pdf.drawString(3.5*cm, altura_pagina - 10.8*cm, str(linea_dir3))
    pdf.drawString(14*cm, altura_pagina - 9.35*cm, 'Fecha: ' + str(fecha_actual))
    pdf.drawString(14*cm, altura_pagina - 10*cm, 'Evento: '+ str(fecha_evento))
    pdf.drawString(14*cm, altura_pagina - 10.65*cm, 'Hasta: '+ str(fecha_evento_hasta))
    l = 60
    pdf.drawString(l, 40, str(color_descrip))
    if b.cobras > 0:
        pdf.drawString(3.5*cm, altura_pagina - 15.25*cm, 'Cobras: '+ str(cobras) + ' UND')
    if b.handsfree > 0:
        pdf.drawString(3.5*cm, altura_pagina - 15.95*cm, 'HandsFree: '+ str(handsfree) + ' UND')
    if b.cascos > 0:
        pdf.drawString(3.5*cm, altura_pagina - 17.45*cm, str('Headset Antiruido'))
        pdf.drawString(14*cm, altura_pagina - 17.45*cm, str(cascos + ' UND'))
    if b.baterias > 0:
        pdf.drawString(3.5*cm, altura_pagina - 19.45*cm, 'Baterias: '+ str(baterias) + ' UND')
    if b.cargadores > 0:
        pdf.drawString(3.5*cm, altura_pagina - 20.15*cm, 'Cargadores: '+ str(cargadores + ' UND'))
    if b.estaciones >0:
        pdf.drawString(3.5*cm, altura_pagina - 16.75*cm, 'HandsFree Tipo Escolta: '+ str(estaciones + ' UND'))
    if b.repetidoras > 0:
        pdf.drawString(3.5*cm, altura_pagina - 20.85*cm, str('Repetidoras'))
        pdf.drawString(14*cm, altura_pagina - 20.85*cm, str(repetidoras + ' UND'))
    #radios
    if rx_amarillas > 0:
        pdf.drawString(3.5*cm, altura_pagina - 11.95*cm, str('EP 450 (Amarillas)'))
        pdf.drawString(14*cm, altura_pagina - 11.95*cm, str(amarillas))
        pdf.drawString(14.5*cm, altura_pagina - 11.95*cm, str(' UND'))
    if rx_moradas > 0:
        pdf.drawString(3.5*cm, altura_pagina - 13.45*cm, str('DEP 450 (Moradas)'))
        pdf.drawString(14*cm, altura_pagina - 13.45*cm, str(moradas))
        pdf.drawString(14.5*cm, altura_pagina - 13.45*cm, str(' UND'))
    if rx_plomo > 0:
        pdf.drawString(3.5*cm, altura_pagina - 12.65*cm, str('DEP 450 (Plomo)'))
        pdf.drawString(14*cm, altura_pagina - 12.65*cm, str(plomo))
        pdf.drawString(14.5*cm, altura_pagina - 12.65*cm, str(' UND'))
    # #y = 110
    font_size = 9
    pdf.setFont("Helvetica", font_size)
    x = 60
    h = 60
    f = 60
    g = 60
    # for i in rx:
    #     pdf.drawString(x, 110, str(i))
    #     #y += 20
    #     x += 30
    cuenta_serial = len(rx)
    if cuenta_serial > 16:
        for indice, i in enumerate(rx): 
            if indice <15:
                pdf.drawString(x, 80, str(i))
                x += 30
            if indice >15 and indice <31:
                pdf.drawString(h, 70, str(i))
                h += 30
            if indice >31 and indice <47:
                pdf.drawString(f, 60, str(i))
                f += 30
            if indice >47 and indice <60:
                pdf.drawString(g, 50, str(i))
                g += 30
    else:
        for i in rx[:cuenta_serial]:
            pdf.drawString(x, 80, str(i))
            #y += 20
            x += 30

    

    pdf.showPage()

    pdf.save()

    buffer.seek(0)

    nombre_archivo = str(id)+'_'+str(fecha_entrega)+'.pdf'
   
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename= {nombre_archivo}'


    return response
@login_required
def buscaOrdenesDevueltas(request):

    if request.method == 'POST':
        nombre = request.POST.get('clienteNombre')

        nombre_cliente = cliente.objects.all()
    
        devueltas = vista_entrada_detalle.objects.filter(cliente = nombre)

        return render(request, 'ordenesDevueltas.html', {"listaOrdenes": devueltas, "lista_cliente":nombre_cliente})
@login_required
def monitor(request):
    fecha_actual = datetime.date.today()

    ordenes = ordenRegistro.objects.filter(estado_id=5, fecha_entrega=fecha_actual)
    ordenes_count = ordenRegistro.objects.filter(estado_id = 5, fecha_entrega = fecha_actual).count()

    ordenes_retiro = ordenRegistro.objects.filter(estado_id=3, fecha_retiro__lte = fecha_actual)
    ordenes_retiro_count = ordenRegistro.objects.filter(estado_id = 3, fecha_retiro__lte = fecha_actual).count()
    return render(request, 'monitor.html', {"listaOrdenesRetiro": ordenes_retiro, "total_ordenes_retiro":ordenes_retiro_count, "listaOrdenes":ordenes, "total_ordenes":ordenes_count})
@login_required
def generarPDFprint(request, id):

    detalle_acce = salidasDetalle.objects.get(id_orden = id)
    b = detalle_acce

    orden = b.id_orden
    salida_id = b.id

    ordenes = ordenRegistro.objects.get(id = orden)

    c = ordenes

    cliente = str(c.cliente)
    radios = str(c.cantidad_radios)
    cobras = str(b.cobras)
    baterias = str(b.baterias)
    cargadores = str(b.cargadores)
    manos_libres = str(b.handsfree)
    cascos = str(b.cascos)
    repetidoras = str(b.repetidoras)
    estaciones = str(b.estaciones)
    fentrega = str(c.fecha_entrega)
    pedido = str(c.id)

    buffer = BytesIO()
    tamano_pagina =  (20*cm, 10*cm)
    pdf = canvas.Canvas(buffer, pagesize=tamano_pagina)
   

    pdf.drawString(2*cm, 9*cm,"N° Pedido: "+ pedido)
    pdf.drawString(2*cm, 8*cm, "Cliente: "+ cliente)
    pdf.drawString(2*cm, 7*cm,"Fecha entrega: " +fentrega)
    if c.cantidad_radios > 0:
        pdf.drawString(2*cm, 6*cm, "Radios: "+ radios)
    if b.cobras > 0:
        pdf.drawString(2*cm, 5.5*cm, "Cobras: "+ cobras)
    if b.baterias > 0:
        pdf.drawString(2*cm, 5*cm, "Baterias extra: "+ baterias)
    if b.cargadores > 0:
        pdf.drawString(2*cm, 4.5*cm, "Cargadores: " + cargadores)
    if b.handsfree > 0:
        pdf.drawString(2*cm, 4*cm,"Manos libres: " + manos_libres)
    if b.cascos > 0:
        pdf.drawString(2*cm, 3.5*cm,"Cascos: " + cascos)
    if b.repetidoras > 0:
        pdf.drawString(2*cm, 3*cm,"Repetidoras: " + repetidoras)
    if b.estaciones > 0:
        pdf.drawString(2*cm, 2.5*cm,"Estaciones: " + estaciones)
    
    
    pdf.showPage()

    pdf.save()

    buffer.seek(0)
    nombre_archivo = str(id)+'_'+str(cliente)+'.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename= {nombre_archivo}'


    return response
@login_required
def generaInformes(request):

    form = formBuscarInformes()
    form_buscar = formBuscarInformes(request.POST)
    if request.method == 'POST':
        if form_buscar.is_valid():
            desde = form_buscar.cleaned_data['fechaBInformeDesde']
            hasta = form_buscar.cleaned_data['fechaBInformeHasta']
            cliente = form_buscar.cleaned_data['cliente']

            result_busqueda = vista_movimiento_radios_tipos.objects.filter(fecha_salida__range=(desde,hasta), cliente = cliente, estado = 'F')
            agrupacion = result_busqueda.values('id_orden', 'fecha_salida').annotate(total_registros=Count('serialrx'))

            agrupacion_por_mes = result_busqueda.annotate(mes=TruncMonth('fecha_salida')).values('mes').annotate(total_registros=Count('serialrx')).values('mes', 'total_registros')


            #########IMAGEN#########

            image1_filename = 'logo_22.jpg'
            image_path1 = f'radios/static/{image1_filename}'

            #generar el PDF 
            buffer = BytesIO()
            pdf = canvas.Canvas(buffer)

            ancho_pagina, altura_pagina = letter = (21.59*cm, 27.94*cm)

            # Calcula el tamaño y posición de la imagen
            image_height = 60
            o = 60
            z = letter[1] - image_height

            image_width01 = 150
            image_height01 = 60

            o1 = 2*cm
            z2 = 27*cm
            
            # Agrega la imagen al documento PDF
            pdf.drawImage(image_path1, o1, z2, width=image_width01, height=image_height01)

            ####TITULO#########
            titulo = "REPORTE SALIDAS RADIOS"
            ancho_texto = pdf.stringWidth(titulo, "Helvetica", 12)
            # Calcular la posición horizontal para centrar
            pos_x = (ancho_pagina - ancho_texto) / 2
            # Definir la posición vertical
            pos_y = altura_pagina - 2*cm         
            
            # ESCRIBIR DATOS DEL PDF ####
            ####TITULO#########
            pdf.setFillColorRGB(0, 0, 1)
            pdf.drawString(pos_x, pos_y, titulo)

            ####NOMBRE DEL CLIENTE#########
            pos_y = altura_pagina - 3*cm
            pdf.setFillColorRGB(1, 0, 0)
            pdf.drawString(3*cm, pos_y, "CLIENTE: "+ str(cliente))
            pos_y = altura_pagina - 3.5*cm
            pdf.drawString(3*cm, pos_y, "Desde: "+ str(desde)+ " - " + "Hasta: " + str(hasta))
            pdf.setFillColorRGB(0, 0, 0)

            ###### TITULOS COLUMNAS############
            x = 4.5*cm
            pdf.drawString(3*cm, altura_pagina - x, "N° ORDEN")
            pdf.drawString(7*cm, altura_pagina - x, "FECHA")
            pdf.drawString(12*cm, altura_pagina - x, "CANTIDAD")


            x = 5.0*cm
            # cliente = str(i.nombre)
            for i in agrupacion:
                cantidad = str(i['total_registros'])
                fecha = str((i['fecha_salida']))
                salida = str(i['id_orden'])
                pdf.drawString(4*cm, altura_pagina - x,salida )
                pdf.drawString(6*cm, altura_pagina - x,fecha )
                pdf.drawString(13*cm, altura_pagina - x,cantidad )
                x += 0.5*cm
            
            x += 1*cm
            resultado_final = agrupacion.aggregate(total_final=Sum('total_registros'))
            total_final = resultado_final['total_final']

            pdf.drawString(12*cm, altura_pagina - x,"____________" )
            x += 1*cm
            pdf.drawString(13*cm, altura_pagina - x,"TOTAL: " + str(total_final) )

            x += 2*cm
            for meses in agrupacion_por_mes:
                result = f"{meses['mes'].strftime('%B')}    =     {meses['total_registros']}"
                pdf.drawString(3*cm, altura_pagina - x, result)
                x += 0.5*cm
            x += 2*cm

            ############GRAFICO BARRAS######

            meses = [item['mes'].strftime('%B') for item in agrupacion_por_mes]
            total_registros_por_mes = [item['total_registros'] for item in agrupacion_por_mes]

            # Crear el gráfico de barras
            ptl.bar(meses, total_registros_por_mes)

            # Personalizar el gráfico (opcional)
            ptl.xlabel('Meses')
            ptl.ylabel('Total de Radios')
            ptl.title('Gráfico de Barras por Mes')
            ptl.xticks(rotation=45)  # Rotar las etiquetas del eje x para mayor legibilidad
            ptl.grid(axis='y', linestyle='--', alpha=0.7)

            buffer1 = BytesIO()
            ptl.savefig(buffer1, format='png')
            buffer1.seek(0)
            image_base641 = base64.b64encode(buffer1.read()).decode()
            grafico = "data:image/png;base64," + image_base641

            x = 10*cm
            pdf.drawImage((grafico), 300, x, width=300, height=250)
            
            pdf.showPage()

            pdf.save()

            buffer.seek(0)
            nombre_archivo = str(cliente)+'.pdf'
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename= {nombre_archivo}'
            return response
    else:
        return render(request, 'informes.html', {'form':form})
@login_required
def generarPDFtotales(request):

    form = formBuscarInformes()

    if request.method == 'POST':
            anio = request.POST.get('years')

            result_busqueda = vista_movimiento_radios_tipos.objects.filter(fecha_salida__range=(
            f'{anio}-01-01',
            f'{anio}-12-31' ), estado = 'F')

            agrupacion_por_mes = result_busqueda.annotate(mes=TruncMonth('fecha_salida')).values('mes').annotate(total_registros=Count('serialrx')).values('mes', 'total_registros')
            agrupacion_por_tipo = result_busqueda.values('tipo').annotate(total_registros=Count('serialrx')).values('tipo', 'total_registros')

            #generar el PDF 
            buffer = BytesIO()
            pdf = canvas.Canvas(buffer)

            ancho_pagina, altura_pagina = letter = (21.59*cm, 27.94*cm)
            image1_filename = 'logo_22.jpg'
            image_path1 = f'radios/static/{image1_filename}'

            image_width01 = 150
            image_height01 = 60

            o1 = 2*cm
            z2 = 27*cm
            
            # Agrega la imagen al documento PDF
            pdf.drawImage(image_path1, o1, z2, width=image_width01, height=image_height01)

           ####TITULO#########
            titulo = "REPORTE SALIDAS RADIOS TOTAL POR MES"
            ancho_texto = pdf.stringWidth(titulo, "Helvetica", 12)
            # Calcular la posición horizontal para centrar
            pos_x = (ancho_pagina - ancho_texto) / 2
            # Definir la posición vertical
            pos_y = altura_pagina - 2*cm         
            
            # ESCRIBIR DATOS DEL PDF ####
            ####TITULO#########
            pdf.setFillColorRGB(0, 0, 1)
            pdf.drawString(pos_x, pos_y, titulo)

            ####NOMBRE DEL CLIENTE#########
            pos_y = altura_pagina - 3*cm
            pdf.setFillColorRGB(1, 0, 0)
            # pdf.drawString(3*cm, pos_y, "CLIENTE: "+ str(cliente))
            pdf.setFillColorRGB(0, 0, 0)

            x = 4.5*cm
            # Extrae los meses y los totales por mes
            for meses in agrupacion_por_mes:
                result = f"{meses['mes'].strftime('%B')}    =     {meses['total_registros']}"
                pdf.drawString(3*cm, altura_pagina - x, result)
                x += 0.5*cm

            x += 2*cm

            ##########TOTAL GENERAL#############

            resultado_final = agrupacion_por_mes.aggregate(total_final=Sum('total_registros'))
            total_final = resultado_final['total_final']

            pdf.drawString(3*cm, altura_pagina - x, "Total año "+ str(anio) + " = " + str(total_final))
            x += 2*cm

             ##########TOTAL GENERAL TIPO#############
            for tipo in agrupacion_por_tipo:
                result_tipo = f"{tipo['tipo']}    =     {tipo['total_registros']}"
                pdf.drawString(3*cm, altura_pagina - x, result_tipo)
                x += 0.5*cm

            x += 2*cm



            ##########GRAFICO MES #################

            etiquetas1 = [mes['mes'].strftime('%B') for mes in agrupacion_por_mes]
            valores1 = [cantidad['total_registros'] for cantidad in agrupacion_por_mes]
            fig1, ax1 = ptl.subplots()
            fig1.set_facecolor('white')
            ax1.pie(valores1, labels=etiquetas1,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'black'})
            ax1.axis('equal')
            ax1.set_title('Distribución por Mes', fontweight='bold', fontdict={'color': 'black', 'fontsize': 16})

            buffer1 = BytesIO()
            ptl.savefig(buffer1, format='png')
            buffer1.seek(0)
            image_base641 = base64.b64encode(buffer1.read()).decode()
            grafico2 = "data:image/png;base64," + image_base641

            x = 18*cm
            pdf.drawImage((grafico2), 300, x, width=250, height=200)

             ##########GRAFICO TIPO #################

            etiquetas2 = [tipo['tipo'] for tipo in agrupacion_por_tipo]
            valores2 = [cantidad['total_registros'] for cantidad in agrupacion_por_tipo]
            fig2, ax2 = ptl.subplots()
            fig2.set_facecolor('white')
            ax2.pie(valores2, labels=etiquetas2,autopct='%1.1f%%',startangle=140, shadow=True, textprops={'color': 'black'})
            ax2.axis('equal')
            ax2.set_title('Distribución por Tipo', fontweight='bold', fontdict={'color': 'black', 'fontsize': 16})

            buffer2 = BytesIO()
            ptl.savefig(buffer2, format='png')
            buffer2.seek(0)
            image_base6412 = base64.b64encode(buffer2.read()).decode()
            grafico3 = "data:image/png;base64," + image_base6412

            x = 10*cm
            pdf.drawImage((grafico3), 300, x, width=250, height=200)

            

            ###########MOSTRAR EL PDF###########
            pdf.showPage()

            pdf.save()

            buffer.seek(0)
            nombre_archivo = ('Totales_Por_Mes_')+ anio +'.pdf'
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename= {nombre_archivo}'
            return response  

            
    else:

        return render(request, 'informes.html', {'form':form})    
@login_required
def auditoria(request):
    return render(request, 'auditoria.html')
@login_required    
def listadocxc(request):

    form = FacturaPDFForm()
    nombre_cliente = cliente.objects.all().order_by('nombre')
    devueltas = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('-fecha_entrega')
    return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def listadocxcfacturado(request):

    form = FacturaPDFForm()
    nombre_cliente = cliente.objects.all().order_by('nombre')
    devueltas = vista_ordenes_cxc.objects.filter(facturado=1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')  # saldo igual a monto_total
).order_by('fecha_entrega')
    return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def listadocxcfacturadoAbono(request):

    form = FacturaPDFForm()
    nombre_cliente = cliente.objects.all().order_by('nombre')
    devueltas = vista_ordenes_cxc.objects.filter(facturado=1, pagado = 0).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo__lt=F('monto_total')
).order_by('fecha_entrega')
    return render(request, 'listadocxcfacturadoabono.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def detalleabonado(request, id):

    encabezado = get_object_or_404(vista_ordenes_cxc, id_orden=id)
    orden = encabezado.id_orden
    salida = encabezado.id_salida
    nombre_cliente = encabezado.cliente
    cliente_ruc = encabezado.ruc
    razon_social = encabezado.razon_social
    ruc_razon_s = encabezado.ruc_razon_social
    monto_factura = encabezado.monto_total
    fecha_evento = encabezado.fecha_evento_desde
    archivo_factura = encabezado.factura_pdf
    cantidad_radios = encabezado.cantidad_radios
    saldo_restante = encabezado.saldo

    abonos = abono_factura.objects.filter(id_orden = id).order_by('fecha_abono')
    suma_montos_abonos = abono_factura.objects.filter(id_orden = id).aggregate(total_abono = Sum('monto_abono'))
    total_abonos_facturas = round(suma_montos_abonos['total_abono'], 2)
    
    return render(request, 'detalleabonos.html', {"lista_cliente":nombre_cliente, 'cliente_ruc':cliente_ruc, 'razon_social':razon_social,
                                                  'ruc_razon_s':ruc_razon_s, 'numero_orden':orden, 'numero_salida':salida,
                                                   'monto_factura': monto_factura, 'fecha_evento': fecha_evento, 'saldo':saldo_restante,
                                                    'lista_abonos':abonos, 'total_abonado':total_abonos_facturas})
@login_required
def listadocxcpagado(request):

    nombre_cliente = cliente.objects.all().order_by('nombre')
    devueltas = vista_ordenes_cxc.objects.filter(pagado = 1).order_by('id_orden')
    return render(request, 'listadocxcpagado.html',{"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente})
@login_required
def subir_factura_pdf(request, id, id_salida):
    registro = get_object_or_404(ordenRegistro, id=id)

    if request.method == 'POST':
        form = FacturaPDFForm(request.POST, request.FILES, instance=registro)
        form_datos_fact = formRegistroMontoFact(request.POST)

        archivo_pdf = request.FILES.get('factura_pdf')
        nombre_archivo = archivo_pdf.name
        # valida que sea un pdf

        if not nombre_archivo.lower().endswith('.pdf'):
                return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <img src="/static/error.png" alt="Error" style="max-width: 200px; margin-bottom: 20px;">
                <h2 style="color: red;">EL ARCHIVO SELECCIONADO NO TIENE UN FORMATO PDF.</h2>
                <h2 style="color: red;">POR FAVOR, VUELVA ATRÁS.</H2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)

        if form.is_valid():
            registro.facturado = True  
            registro.save()  
            form.save(commit=False)  
            form.instance = registro  
            form.save()  
            
        
        if form_datos_fact.is_valid():
            monto_base = form_datos_fact.cleaned_data['monto_base']
            igv = form_datos_fact.cleaned_data['igv']
            monto_total = form_datos_fact.cleaned_data['monto_total']
            detraccion = form_datos_fact.cleaned_data['detraccion']
            porcentaje_detrac = request.POST.get('porc_detraccion')
            if porcentaje_detrac:
                porcentaje_detrac = float(porcentaje_detrac)
            sunat = True
            if not detraccion:
                detraccion = 0
                porcentaje_detrac = 0
            guarda_datos_fact = contable(monto_base = monto_base, igv = igv, monto_total = monto_total, id_salida = id_salida,
                                         id_orden = id, saldo = monto_total, porcentaje_detrac = porcentaje_detrac, detraccion = detraccion,
                                         sunat = sunat)
            
            try:
                guarda_datos_fact.save()
            except Exception as e:
                error_message = f"Error al guardar los datos: {e}"
                return HttpResponse(error_message) 
            # guarda_datos_fact.save()

            return redirect('/listadocxcfacturado/')  # Redirige a la vista 

    else:
        form = FacturaPDFForm(instance=registro)

    return redirect('/listadocxcfacturado/')
@login_required
def subir_factura_pdf_no_sunat(request, id, id_salida):
    registro = get_object_or_404(ordenRegistro, id=id)

    if request.method == 'POST':
        
        form_datos_fact =  formRegistroMontoFactNoSunat(request.POST)

        if form_datos_fact.is_valid():
            registro.facturado = True  
            registro.save()  
        
        if form_datos_fact.is_valid():
            monto_base = 0
            igv = 0
            monto_total = form_datos_fact.cleaned_data['monto_total']
            detraccion = 0
            porcentaje_detrac = 0
            sunat = 0
            if not detraccion:
                detraccion = 0
                porcentaje_detrac = 0
            guarda_datos_fact = contable(monto_base = monto_base, igv = igv, monto_total = monto_total, id_salida = id_salida,
                                         id_orden = id, saldo = monto_total, porcentaje_detrac = porcentaje_detrac, detraccion = detraccion,
                                         sunat = sunat)
            guarda_datos_fact.save()
            return redirect('/listadocxcfacturado/')  # Redirige a la vista 

    else:

        return redirect('/listadocxcfacturado/')
@login_required
def ver_factura_pdf(request, id):
    registro = get_object_or_404(ordenRegistro, id=id)
    
    if registro.factura_pdf:
        # Crea una respuesta HTTP
        response = HttpResponse(
            registro.factura_pdf,
            content_type='application/pdf'
        )
        # Establece el encabezado para la descarga
        response['Content-Disposition'] = f'attachment; filename="{registro.factura_pdf.name}"'
        return response
    else:
        # Manejo si no hay PDF disponible
        return HttpResponse("No hay archivo PDF disponible para descargar.")
@login_required
def ver_comprobante_pago(request, id):
    registro = get_object_or_404(contable, id_orden=id)
    
    if registro.comprobante_pago:
        # Crea una respuesta HTTP
        response = HttpResponse(
            registro.comprobante_pago,
            content_type='image/jpeg'
        )
        # Establece el encabezado para la descarga
        response['Content-Disposition'] = f'attachment; filename="{registro.comprobante_pago.name}"'
        return response
    else:
        # Manejo si no hay PDF disponible
        return HttpResponse("No hay archivo disponible para descargar.")
@login_required
def ver_comprobante_pago_abono(request, id):
    registro = get_object_or_404(abono_factura, id=id)
    
    if registro.comprobante_pago:
        # Crea una respuesta HTTP
        response = HttpResponse(
            registro.comprobante_pago,
            content_type='image/jpeg'
        )
        # Establece el encabezado para la descarga
        response['Content-Disposition'] = f'attachment; filename="{registro.comprobante_pago.name}"'
        return response
    else:
        # Manejo si no hay PDF disponible
        return HttpResponse("No hay archivo disponible para descargar.")
@login_required    
def form_registra_fact(request, id, cliente, ruc, id_salida, razon_social):

    if ruc == 'None':
        print(ruc)
        return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <img src="/static/error.png" alt="Error" style="max-width: 200px; margin-bottom: 20px;">
                <h2 style="color: red;">DEBE INGRESAR UNA RAZÓN SOCIAL PARA PODER AGREGAR UNA FACTURA</h2>
                <h2 style="color: red;">POR FAVOR VUELVA ATRÁS.</H2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)
    else:
        form = FacturaPDFForm()
        form_datos_fact = formRegistroMontoFact()
        return render(request, 'formsubirfact.html', {'id':id, 'cliente':cliente,'ruc':ruc,'id_salida':id_salida,'form':form, 'formDatosFact':form_datos_fact,
                                                      'razon_social':razon_social})
@login_required
def form_registra_fact_no_sunat(request, id, cliente, ruc, id_salida, razon_social):

    form_datos_fact = formRegistroMontoFactNoSunat()
    
    return render(request, 'formsubirfactNosunat.html', {'id':id, 'cliente':cliente,'ruc':ruc,'id_salida':id_salida,'formDatosFact':form_datos_fact,
                                                         'razon_social':razon_social})
@login_required
def revertir_factura(request, id):
    
    registro = get_object_or_404(ordenRegistro, id=id)
    registro.facturado = False
    registro.factura_pdf = ''
    registro.save()

    registro_datos = get_object_or_404(contable, id_orden=id)
    registro_datos.delete()

    if abono_factura.objects.filter(id_orden = id).exists():
        abono_factura.objects.filter(id_orden = id).delete()

    return redirect('/listadocxc/')
@login_required
def revertir_abono(request, id):
    
    registro_datos = get_object_or_404(abono_factura, id=id)
    monto_abono = registro_datos.monto_abono
    id_orden = registro_datos.id_orden
    registro_datos_contable = get_object_or_404(contable, id_orden=id_orden)
    registro_datos_contable.saldo = registro_datos_contable.saldo + monto_abono
    registro_datos_contable.save()
    registro_datos.delete()

    #buscar el saldo en contable y actualizar 
    return redirect('/listadocxcfacturadoabono/')
@login_required
def form_registrar_pago(request, id, cliente, ruc, id_salida, razon_social):

    form_datos_pago = formRegistroMontopago()
    form_comprobante_pago = comprobantePagoForm()
    saldo_factura = get_object_or_404(contable, id_orden=id)
    saldo = saldo_factura.saldo
    monto_base = saldo_factura.monto_base
    igv = saldo_factura.igv
    monto_total = saldo_factura.monto_total
    monto_detraccion = saldo_factura.detraccion

    if abono_factura.objects.filter(id_orden = id).exists():
                    suma_montos_abonos = abono_factura.objects.filter(id_orden = id).aggregate(total_abono = Sum('monto_abono'))
                    total_abonos_facturas = suma_montos_abonos['total_abono']

    else:
        total_abonos_facturas = 0

    
    return render(request, 'formregistrapago.html',{'id':id, 'cliente':cliente,'ruc':ruc,'id_salida':id_salida,
                                                    'form_datos_pago':form_datos_pago, 'saldo':saldo, 'form_comprobante_pago':form_comprobante_pago,
                                                    'monto_base':monto_base, 'igv':igv, 'monto_total':monto_total, 
                                                    'total_abonos_facturas':total_abonos_facturas, 'monto_detraccion':monto_detraccion,
                                                    'razon_social':razon_social})
@login_required
def registrar_pago(request, id, id_salida):

    # validar si el monto abonado es igual al monto total de la factura, es pago total y va directo a contable. 
    # si es menor debe verificar los abonos y comparar contra el  va contra la tabla abono y si es mayor da la alerta.
    
    registro = get_object_or_404(contable, id_orden=id)

    if request.method == 'POST':
        
        form_datos_pago = formRegistroMontopago(request.POST)
        form = comprobantePagoForm(request.POST, request.FILES, instance=registro)
            
        if form_datos_pago.is_valid():
            abono = form_datos_pago.cleaned_data['abono']
            fecha_pago = form_datos_pago.cleaned_data['fecha_pago']
            referencia_pago = form_datos_pago.cleaned_data['referencia_pago']
            nuevo_saldo = registro.saldo - abono
            
            # si monto es mayor
            if nuevo_saldo < 0:
                return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <img src="/static/error.png" alt="Error" style="max-width: 200px; margin-bottom: 20px;">
                <h2 style="color: red;">EL MONTO INGRESADO ES MAYOR AL MONTO TOTAL DE LA FACTURA.</h2>
                <h2 style="color: red;">POR FAVOR VUELVE ATRÁS Y CORRIJE.</H2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)
            # si es pago total
            if nuevo_saldo == 0:
                if abono_factura.objects.filter(id_orden = id).exists():
                    id_orden = id
                    id_salida = id_salida
                    monto_abono = form_datos_pago.cleaned_data['abono']
                    fecha_abono = form_datos_pago.cleaned_data['fecha_pago']
                    referencia_abono = form_datos_pago.cleaned_data['referencia_pago']
                    mensaje = "PAGADO EN ABONOS"

                    # VERIFICAR SI LA SUMATORIA DE LOS ABONOS NO SOBREPASA EL MONTO DE FACTURA

                    registro = get_object_or_404(contable, id_orden=id)
                    monto_total_factura = registro.monto_total

                    if abono_factura.objects.filter(id_orden = id_orden).exists():
                        suma_montos_abonos = abono_factura.objects.filter(id_orden = id_orden).aggregate(total_abono = Sum('monto_abono'))
                        total_abonos_facturas = suma_montos_abonos['total_abono']
                    else:
                        total_abonos_facturas = 0
                
                    if (total_abonos_facturas + monto_abono) > monto_total_factura:
                        return HttpResponse(f"""
                    <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                    <img src="/static/error.png" alt="Error" style="max-width: 200px; margin-bottom: 20px;">
                    <h2 style="color: red;">LA SUMATORIA DE LOS ABONOS ES MAYOR AL MONTO TOTAL DE LA FACTURA.</h2>
                    <h2 style="color: red;">POR FAVOR VUELVE ATRÁS Y CORRIJE.</H2>
                    <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                    Volver atrás
                    </button>
                    </div>
                    """)
                    guarda_datos_abono = abono_factura(id_orden = id_orden, id_salida = id_salida, monto_abono = monto_abono,
                                                   fecha_abono = fecha_abono, referencia_abono = referencia_abono)
                    guarda_datos_abono.save()

                    registro = get_object_or_404(contable, id_orden=id)
                    nuevo_saldo = registro.saldo - abono
                    registro.saldo = nuevo_saldo
                    registro.pagado = True
                    registro.fecha_pago = fecha_abono
                    registro.referencia_pago = referencia_abono  
                    registro.save()
                # registro = get_object_or_404(abono_factura, id_orden=id_orden)
                    form = comprobanteabonoForm(request.POST, request.FILES, instance=guarda_datos_abono)
                    if form.is_valid():
                # form.instance = registro
                        form.save()
                #AQUI UN RETURN AL TEMPLATE CON EL MENSAJE 
                else:
                    form = comprobantePagoForm(request.POST, request.FILES, instance=registro)
                    monto_total_factura = registro.monto_total
                    
                    if abono == monto_total_factura:
                        registro.abono = abono
                    else:
                        registro.abono = 0
                    
                    registro.fecha_pago = fecha_pago
                    registro.referencia_pago = referencia_pago
                    registro.saldo = nuevo_saldo 
                    registro.save()
                
                    if form.is_valid():
                        registro.pagado = True  
                        registro.save()  # Guarda los cambios en el registro
                        form.save(commit=False)  # Prepara el formulario para guardar
                        form.instance = registro  # Asigna la instancia actualizada al formulario
                        form.save()

            # si es un abono va contra la tabla abono
            if nuevo_saldo > 0:
                id_orden = id
                id_salida = id_salida
                monto_abono = form_datos_pago.cleaned_data['abono']
                fecha_abono = form_datos_pago.cleaned_data['fecha_pago']
                referencia_abono = form_datos_pago.cleaned_data['referencia_pago']

                # VERIFICAR SI LA SUMATORIA DE LOS ABONOS NO SOBREPASA EL MONTO DE FACTURA

                registro = get_object_or_404(contable, id_orden=id)
                monto_total_factura = registro.monto_total

                if abono_factura.objects.filter(id_orden = id_orden).exists():
                    suma_montos_abonos = abono_factura.objects.filter(id_orden = id_orden).aggregate(total_abono = Sum('monto_abono'))
                    total_abonos_facturas = suma_montos_abonos['total_abono']
                else:
                    total_abonos_facturas = 0
                
                if (total_abonos_facturas + monto_abono) > monto_total_factura:
                    return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <img src="/static/error.png" alt="Error" style="max-width: 200px; margin-bottom: 20px;">
                <h2 style="color: red;">LA SUMATORIA DE LOS ABONOS ES MAYOR AL MONTO TOTAL DE LA FACTURA.</h2>
                <h2 style="color: red;">POR FAVOR VUELVE ATRÁS Y CORRIJE.</H2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)

                guarda_datos_abono = abono_factura(id_orden = id_orden, id_salida = id_salida, monto_abono = monto_abono,
                                                   fecha_abono = fecha_abono, referencia_abono = referencia_abono)
                guarda_datos_abono.save()

                registro = get_object_or_404(contable, id_orden=id)
                nuevo_saldo = registro.saldo - abono
                registro.saldo = nuevo_saldo
                registro.save()

                # registro = get_object_or_404(abono_factura, id_orden=id_orden)
                form = comprobanteabonoForm(request.POST, request.FILES, instance=guarda_datos_abono)
                if form.is_valid():
                # form.instance = registro
                    form.save()
            return redirect('/listadocxcfacturadoabono/')
                
                # tengo que ajustar el saldo en contable
            
        return redirect('/listadopagado/')  # Redirige a la vista que quieras después de guardar

    else:
        return redirect('/listadopagado/')
@login_required
def reportescxc(request):

    #tab1- General - solo prueba
    por_emitir = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('id_orden').count()

    emitidas_por_cobrar = vista_ordenes_cxc.objects.filter(facturado=1).filter(Q(referencia__isnull=True) | Q(referencia='')).order_by('id_orden').count()
    total_saldo = vista_ordenes_cxc.objects.filter(facturado=1).filter(Q(referencia__isnull=True) | Q(referencia='')).aggregate(Sum('saldo'))
    total_saldo_value = total_saldo['saldo__sum']

    cobradas = vista_ordenes_cxc.objects.filter(pagado = 1).order_by('id_orden').count()
    total_cobrado = vista_ordenes_cxc.objects.filter(pagado=1).aggregate(Sum('abono'))
    total_cobrado_value = total_cobrado['abono__sum']

    form = FacturaPDFForm()
    nombre_cliente = cliente.objects.all().order_by('nombre')
    devueltas = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('id_orden')

    #tab2- General - solo prueba

    if request.method == 'POST':
        nombre = request.POST.get('clienteNombre')
           
        por_emitir_detalle = vista_ordenes_cxc.objects.filter(facturado = 0, cliente = nombre).order_by('id_orden').count()

        emitidas_por_cobrar_detalle = vista_ordenes_cxc.objects.filter(facturado=1, cliente = nombre, pagado = 0).order_by('id_orden').count()
        total_saldo_detalle = vista_ordenes_cxc.objects.filter(facturado=1, cliente = nombre, pagado = 0).aggregate(Sum('saldo'))
        total_saldo_value_detalle = total_saldo_detalle['saldo__sum']

        cobradas_detalle = vista_ordenes_cxc.objects.filter(pagado = 1, cliente = nombre).order_by('id_orden').count()
        total_cobrado_detalle = vista_ordenes_cxc.objects.filter(pagado=1, cliente = nombre).aggregate(Sum('abono'))
        total_cobrado_value_detalle = total_cobrado_detalle['abono__sum']
    else:
        nombre = ""
        por_emitir_detalle = 0
        emitidas_por_cobrar_detalle = 0
        total_saldo_value_detalle = 0.00
        total_cobrado_value_detalle = 0.00
        cobradas_detalle = 0

    return render(request, 'reportescxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form,
                                                 'emitir':por_emitir, 'por_cobrar':emitidas_por_cobrar, 'cobradas':cobradas , 'saldo':total_saldo_value,
                                                 'total_cobrado':total_cobrado_value, 'nombre_cliente':nombre, 'emitir_detalle':por_emitir_detalle,
                                                 'por_cobrar_detalle':emitidas_por_cobrar_detalle, 'saldo_detalle': total_saldo_value_detalle,
                                                 'total_cobrado_detalle':total_cobrado_value_detalle, 'cobradas_detalle':cobradas_detalle })
@login_required
def buscarlistadocxccliente(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        nombre = request.POST.get('clienteNombre')

        if nombre == " ":
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(cliente = nombre, facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocxcruc(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        ruc = request.POST.get('ruc_cliente').strip()

        if not ruc:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(ruc_razon_social = ruc, facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocxcidsalida(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        id_salida = request.POST.get('id_salida').strip()

        if not id_salida:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(id_salida = id_salida, facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocxcidorden(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        id_orden = request.POST.get('id_orden').strip()

        if not id_orden:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(id_orden = id_orden, facturado = 0).order_by('fecha_entrega')
            return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxc.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadofacturadocliente(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        nombre = request.POST.get('clienteNombre')

        if nombre == " ":
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(cliente = nombre, facturado = 1).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadofacturadoruc(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        ruc = request.POST.get('ruc_cliente').strip()

        if not ruc:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(ruc_razon_social = ruc, facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadofacturadoidsalida(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        id_salida = request.POST.get('id_salida').strip()

        if not id_salida:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(id_salida = id_salida, facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadofacturadoidorden(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        id_orden = request.POST.get('id_orden').strip()

        if not id_orden:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(id_orden = id_orden, facturado = 1).filter(Q(referencia__isnull=True) | Q(referencia='')).filter(
    saldo=F('monto_total')).order_by('fecha_evento_desde')
            return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcfacturado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocobradocliente(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        nombre = request.POST.get('clienteNombre')

        if nombre == " ":
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(pagado = 1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(cliente = nombre, pagado =1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocobradoruc(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        ruc = request.POST.get('ruc_cliente').strip()

        if not ruc:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(pagado = 1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(ruc_razon_social = ruc, pagado = 1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocobradoidsalida(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        id_salida = request.POST.get('id_salida').strip()

        if not id_salida:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(pagado = 1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(id_salida = id_salida, pagado = 1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def buscarlistadocobradoreferencia(request):

    form = FacturaPDFForm()

    if request.method == 'POST':
        referencia = request.POST.get('referencia').strip()

        if not referencia:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(pagado = 1).order_by('id_orden').order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        else:
            nombre_cliente = cliente.objects.all().order_by('nombre')
            devueltas = vista_ordenes_cxc.objects.filter(referencia = referencia, pagado = 1).order_by('fecha_evento_desde')
            return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
        
    return render(request, 'listadocxcpagado.html', {"listaOrdenescerradas": devueltas, "lista_cliente":nombre_cliente, 'form': form})
@login_required
def pdfporcobrar_detalle(request, cliente):
    try:
        if vista_ordenes_cxc.objects.filter(cliente = cliente, facturado = 1, pagado = 0).order_by('id_orden').exists():
            
            devueltas = vista_ordenes_cxc.objects.filter(cliente = cliente, facturado = 1, pagado = 0).order_by('id_orden')
            total_saldo_detalle = vista_ordenes_cxc.objects.filter(facturado=1, cliente = cliente, pagado = 0).aggregate(Sum('saldo'))
            total_saldo_value_detalle = total_saldo_detalle['saldo__sum']
            total_saldo_value_detalle = format(total_saldo_value_detalle, ".2f")
            ruc_unico = devueltas.values('ruc').distinct().first()
            ruc_valor = ruc_unico['ruc']

            buffer = BytesIO()
            tamano_pagina =  (21.59*cm, 27.94*cm)
            pdf = canvas.Canvas(buffer, pagesize=tamano_pagina)

            image_filename = 'logo_22.jpg'
            image_path = f'radios/static/{image_filename}'

            # Calcula el tamaño y posición de la imagen
            image_width = 60
            image_height = 60
            o = 60
            z = letter[1] - image_height

            image_width01 = 250
            image_height01 = 70

            o1 = 60
            z2 = z - image_height -20
            
            pdf.drawImage(image_path, o1, z2, width=image_width01, height=image_height01)

            font_size = 14
            pdf.setFont("Helvetica", font_size)
            pdf.setFillColorRGB(0, 0, 1)
            pdf.drawString(2*cm, 22*cm, "RELACIÓN DE FACTURAS PENDIENTE POR COBRAR")
            pdf.setFillColorRGB(0, 0, 0)
            pdf.setFont("Helvetica-Bold", font_size)
            pdf.drawString(2*cm, 21*cm, "CLIENTE: "+ cliente)
            if ruc_valor is None:
                ruc_valor = ""
                pdf.drawString(2*cm, 20.5*cm, "RUC: "+ ruc_valor)
            else:
                pdf.drawString(2*cm, 20.5*cm, "RUC: "+ ruc_valor)
            pdf.setFillColorRGB(0, 0, 0)
            pdf.setFont("Helvetica", font_size)

            font_size = 12
            pdf.setFont("Helvetica", font_size)

            h = 2*cm
            v = 19*cm

            for i in devueltas:
                if abono_factura.objects.filter(id_orden = i.id_orden).exists():
                    abonos = abono_factura.objects.filter(id_orden = i.id_orden).aggregate(Sum('monto_abono'))
                    total_abono_value = abonos['monto_abono__sum']
                    total_abono_value = format(total_abono_value, ".2f")
                else:
                    total_abono_value = 0.00
                h = 2*cm
                pdf.setFillColorRGB(0, 0, 1)
                pdf.drawString(h, v, 'Fecha Evento:')
                h += 3*cm
                pdf.setFillColorRGB(0, 0, 0)
                fecha_evento = str(i.fecha_evento_desde) 
                pdf.drawString(h, v, fecha_evento)
                v -= 0.5*cm
                h = 2*cm
                pdf.drawString(h, v, 'N° Orden:')
                salida = str(i.id_salida)
                h += 3*cm
                pdf.drawString(h, v, salida)
                h += 6*cm
                pdf.drawString(h, v, 'Monto Total Factura:')
                h += 5*cm
                monto_factura = str(i.monto_total)
                pdf.drawString(h, v, 'S/' + ' ' + monto_factura)
                v -= 0.5*cm
                h = 2*cm
                pdf.drawString(h, v, 'Cantidad Radios:')
                h += 4*cm
                radios_cantidad = str(i.cantidad_radios)
                pdf.drawString(h, v, radios_cantidad)
                h += 5*cm
                pdf.drawString(h, v, 'Monto Total Abonado:')
                h +=5*cm
                pdf.drawString(h,v, 'S/' + ' ' + str(total_abono_value))
                h = 2*cm
                v -= 0.5*cm
                if i.razon_social is not None:
                    pdf.drawString(h, v, 'Razon Social:')
                    h += 4*cm
                    razon_social = str(i.razon_social)
                    pdf.drawString(h, v, razon_social)
                    h = 2*cm
                    v -= 0.5*cm
                    pdf.drawString(h, v, 'RUC:')
                    h += 4*cm
                    ruc_razon_social = str(i.ruc_razon_social)
                    pdf.drawString(h, v, ruc_razon_social)

                v -= 1*cm

                #colocar Razon social 
            
            v -=2*cm
            h = 2*cm
            pdf.setFillColorRGB(1, 0, 0)
            font_size = 14
            pdf.setFont("Helvetica-Bold", font_size)
            pdf.drawString(h, v, 'Total por Cobrar:')
            h = 6.5*cm
            pdf.drawString(h, v, 'S/' + ' ' + str(total_saldo_value_detalle))
            pdf.setFillColorRGB(0, 0, 0)
            
            pdf.showPage()
            pdf.save()

            buffer.seek(0)
            nombre_archivo = str(cliente)+'.pdf'
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename= {nombre_archivo}'

            return response
    except Exception as e:
                error_message = f"Error al guardar los datos: {e}"
                return HttpResponse(error_message) 

    else:
        return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <h2 style="color: red;">EL CLIENTE NO TIENE FACTURAS PENDIENTE DE PAGO</h2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)

@login_required
def controlrxevento(request, id):

    try:
        form_serial_recojo = rxcontroleventoformRecojo()
        if controlrx_event_dia.objects.filter(id_salida = id, activo = True).exists():
            busca_dia = controlrx_event_dia.objects.get(id_salida = id, activo = True)
            dia_entrega = busca_dia.dia

        # contar radios disponibles en la orden
        
        radios_en_orden = movimientoRadios.objects.filter(id_salida = id, estado = 'F').count()

        # Si se ha presionado "Limpiar Todo", vaciamos la sesión
        if request.GET.get('limpiar', False):
            request.session.pop('responsable', None)
            request.session.pop('telefono', None)
            request.session.pop('seriales', None)
            return redirect('controlrxevento', id=id)

        # Inicializamos el formulario del responsable solo si no está en sesión
        if 'responsable' not in request.session:
            request.session['seriales'] = []  # Inicializamos la lista de seriales

        # Procesamos el formulario cuando el usuario hace POST
        if request.method == 'POST':
            # Si aún no hemos guardado el responsable, guardamos el valor de la sesión
            if 'responsable' not in request.session:
                form_responsable = ResponsableForm(request.POST)
                form_serial = rxcontroleventoform()  # Inicializamos formulario de serial vacío
                if form_responsable.is_valid():
                    # Guardamos el responsable en la sesión
                    request.session['responsable'] = form_responsable.cleaned_data['responsable']
                    # request.session['telefono'] = form_responsable.cleaned_data['telefono']
                    tel_modificado = form_responsable.cleaned_data['telefono']
                    request.session['telefono'] = re.sub(r'[^\d]+', '', tel_modificado)

                    request.session.modified = True
                    return redirect('controlrxevento', id=id)
            else:
                # Si el responsable ya está en sesión, procesamos los seriales
                form_responsable = ResponsableForm()  # Formulario de responsable vacío porque ya lo tenemos
                form_serial = rxcontroleventoform(request.POST)
                if form_serial.is_valid():


                    ###### INICIO VALIDACIONES #######
                        # QUE EL SERIAL PERTENEZCA A LA ORDEN
                    valida_serial = form_serial.cleaned_data['serial']
                    if not controlrxevent.objects.filter(serial = valida_serial, estadorx = 'S').exists():
                        return HttpResponse(f"""
                            <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                            <h3 style="color: red;">EL SERIAL INTRODUCIDO NO PERTENECE A ESTA ORDEN</h3>
                            <h3 style="color: red;">POR FAVOR VUELVE ATRÁS Y VERIFICA.</H3>
                            <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                            Volver atrás
                            </button>
                            </div>
                            """)
                        # QUE NO SE HAYA ENTREGADO SIN SER DEVUELTO
                    if controlrxevent.objects.filter(serial = valida_serial, estadorx = 'E', id_salida = id).exists():
                        return HttpResponse(f"""
                            <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                            <h3  style="color: red;">EL ESTADO DE ESTE SERIAL ES "ENTREGADO".</h3>
                            <h3  style="color: red;">POR FAVOR VUELVE ATRÁS Y VERIFICA.</H3>
                            <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                            Volver atrás
                            </button>
                            </div>
                            """)
                        # QUE NO SE REPITA EN EL SESSION
                    if 'seriales' in request.session and any(item['serial'] == valida_serial for item in request.session['seriales']):
                    # El serial ya existe en la sesión
                        return HttpResponse(f"""
                            <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                                <h2 style="color: red;">EL SERIAL YA SE ENCUENTRA AGREGADO</h2>
                                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                                    Volver atrás
                                </button>
                            </div>
                        """)

                    # Agregamos el serial a la lista en la sesión.
                    request.session['seriales'].append({
                        'serial': form_serial.cleaned_data['serial'],
                        'responsable': request.session['responsable'],
                        'telefono': request.session['telefono'],
                        'estado': 'S'  # Pendiente
                    })
                    request.session.modified = True
                    return redirect('controlrxevento', id=id)

        else:
            # Formularios en GET
            form_responsable = ResponsableForm()
            form_serial = rxcontroleventoform()

        # Guardamos todos los seriales al hacer clic en "Guardar Todos los Seriales"
        if request.GET.get('guardar', False):
            
            hora_actual = d.now()
            hora_actual = hora_actual.strftime("%H:%M")
            usuario_bk = request.user.username
            
            for item in request.session['seriales']:
                # Guardamos cada serial en la base de datos
                controlrxevent.objects.create(
                    id_salida=id,
                    serial=item['serial'],
                    responsable=item['responsable'],
                    telefono = item['telefono'],
                    estadorx='E',
                    hora_entrega = hora_actual,
                    dia = dia_entrega,
                    responsable_bk = usuario_bk
                    ####### METER EL DIA AQUI ###########    
                )
            # Pasamos los datos de los seriales y responsable a la sesión para impresión
            responsable = request.session.get('responsable', '')
            telefono = request.session.get('telefono', '')
            seriales = request.session.get('seriales', [])

            # Guardamos en la sesión para que esté disponible en el template de impresión
            request.session['responsable_print'] = responsable
            request.session['seriales_print'] = seriales
            request.session['telefono_print'] = telefono
            request.session.modified = True

            # Limpiamos la lista de seriales y el responsable después de guardar
            del request.session['seriales']
            del request.session['responsable']
            del request.session['telefono']
            return redirect('print_template', id = id)
            # return redirect('controlrxevento', id=id)
            
        ######### DATOS PARA TAB 3 ##################

        total_entregado = controlrxevent.objects.filter(id_salida = id, estadorx = 'E').count()
        total_disponible = radios_en_orden - total_entregado
        total_recogido = controlrxevent.objects.filter(id_salida = id, estadorx = 'R').count()
        tabla_resumen = controlrxevent.objects.filter(Q(estadorx='E') | Q(estadorx='R'), id_salida=id).order_by('-dia','responsable', 'hora')

        # Obtenemos los seriales agregados temporalmente
        seriales_temporales = request.session.get('seriales', [])
        return render(request, 'controlrxeventos.html', {
            'form_responsable': form_responsable,
            'form_serial': form_serial, 
            'seriales_temporales': seriales_temporales,
            'id': id,
            'responsable': request.session.get('responsable'),
            'telefono': request.session.get('telefono'),
            'radios_orden': radios_en_orden,
            'form_recojo':form_serial_recojo,
            'total_entregado':total_entregado,
            'total_disponibles': total_disponible,
            'total_recogido': total_recogido,
            'tabla_resumen':tabla_resumen,
            'dia':dia_entrega
        })
    except Exception as e:
                error_message = f"Error al leer los datos: {e}"
                return HttpResponse(error_message) 

@login_required
def cargarxordenevento(request, id):

    try:

        rx_orden = movimientoRadios.objects.filter(id_salida = id, estado = 'F')
        registros = [
            controlrxevent(
                id_salida = i.id_salida,
                serial = i.serial,
                responsable = "",
                estadorx = 'S'
            )

            for i in rx_orden
        ]

        if not controlrxevent.objects.filter(id_salida = id, estadorx = 'S').exists():
            controlrxevent.objects.bulk_create(registros)
            #### ESCRIBE EN LA TABLA DE CONTROL DE DIA ####
            if not controlrx_event_dia.objects.filter(id_salida = id).exists():
                guardar_control_evento = controlrx_event_dia(id_salida = id, dia = 1, activo = True)
                guardar_control_evento.save()
            
        return redirect('controlrxevento', id=id)
    except Exception as e:
                error_message = f"Error al leer los datos: {e}"
                return HttpResponse(error_message)  

@login_required
def controlrxeventorecojo(request, id):

    if request.method == 'POST':
        form_recojo = rxcontroleventoformRecojo(request.POST)

        if form_recojo.is_valid():
            serial = form_recojo.cleaned_data['serial_recojo']

            # calcular hora suma 3 de diferencia con Oregon

            hora_actual = d.now()
            # hora_actual = hora_actual + timedelta(hours=3)
            hora_actual = hora_actual.strftime("%H:%M")
            busca_dia = controlrx_event_dia.objects.get(id_salida = id, activo = True)
            dia_recojo = busca_dia.dia
            usuario_bk = request.user.username

            

            # if controlrxevent.objects.filter(serial = serial, id_salida = id, estadorx = 'R', hora__isnull = False).exists():
            #     return HttpResponse(f"""
            #             <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
            #                 <h2 style="color: red;">ESTE SERIAL YA FUE RECOGIDO</h2>
            #                 <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
            #                     Volver atrás
            #                 </button>
            #             </div>
    
            #         """)
            
            try:
                radio_recogida = controlrxevent.objects.get(serial = serial, id_salida = id, estadorx = 'E')
                radio_recogida.estadorx = 'R'
                radio_recogida.hora = hora_actual
                radio_recogida.dia = dia_recojo
                radio_recogida.responsable_bk = usuario_bk
                radio_recogida.save()
                
                return redirect('controlrxevento', id=id)
            except controlrxevent.DoesNotExist:
                return HttpResponse(f"""
                        <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                            <h2 style="color: red;">ESTE SERIAL NO HA SIDO ENTREGADO</h2>
                            <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                                Volver atrás
                            </button>
                        </div>
    
                    """)
    else:
        return redirect('controlrxevento', id=id)
@login_required
def cerrar_dia_rx_evento(request, id):

    try:

        if controlrx_event_dia.objects.filter(id_salida = id).exists():
            dia_anterior = controlrx_event_dia.objects.get(id_salida = id, activo = True)
            dia_anterior.activo = False
            dia_anterior.save()

            
            dia_nuevo = dia_anterior.dia + 1

            nuevo_dia = controlrx_event_dia(id_salida = id, dia = dia_nuevo, activo = True)
            nuevo_dia.save()

            ############### CREAR UN ESPEJO ###############

            dia_respaldo = dia_anterior.dia

            datos_inicial = controlrxevent.objects.filter(id_salida = id, dia = dia_respaldo)
            total_rx_usados = controlrxevent.objects.filter(id_salida = id, dia = dia_respaldo).count()
            total_rx_por_recojo = controlrxevent.objects.filter(id_salida = id, dia = dia_respaldo, estadorx = 'E')
            datos_espejo = []
            
            for i in datos_inicial:
                dato_espejo = espejo_dia_control_rx(id_salida = i.id_salida,
                                                    serial = i.serial,
                                                    responsable = i.responsable, 
                                                    estadorx = i.estadorx,
                                                    telefono = i.telefono,
                                                    hora = i.hora,
                                                    hora_entrega = i.hora_entrega,
                                                    dia = i.dia)
                datos_espejo.append(dato_espejo)
            espejo_dia_control_rx.objects.bulk_create(datos_espejo)

            ############# GENERAR UN EXCEL #################

            rpt_dia = espejo_dia_control_rx.objects.filter(id_salida = id, dia = dia_respaldo)
            wb = openpyxl.Workbook()
            hoja = wb.active

            hoja['A1'] = 'ID SALIDA'
            hoja['B1'] = 'RESPONSABLE'
            hoja['C1'] = 'TELÉFONO'
            hoja['D1'] = 'SERIAL'
            hoja['E1'] = 'TIPO MOVIMIENTO'
            hoja['F1'] = 'HORA ENTREGA'
            hoja['G1'] = 'HORA RECOJO'
            hoja['H1'] = 'DÍA'

            row = 2

            for i in rpt_dia:
                hoja.cell(row, 1, i.id_salida)
                hoja.cell(row, 2, i.responsable)
                hoja.cell(row, 3, i.telefono)
                hoja.cell(row, 4, i.serial)
                hoja.cell(row, 5, i.estadorx)
                hoja.cell(row, 6, i.hora_entrega)
                hoja.cell(row, 7, i.hora)
                hoja.cell(row, 8, i.dia)
                row += 1

            row += 3
            hoja.cell(row, 1, f'TOTAL RADIOS USADAS DEL DÍA: {total_rx_usados}')
            
            row += 2
            hoja.cell(row, 1, f'RADIOS PENDIENTE POR RECOGER DÍA {dia_respaldo}')
            row += 1

            hoja.cell(row, 1, 'RESPONSABLE')
            hoja.cell(row, 2, 'SERIAL')
            hoja.cell(row, 3, 'HORA DE ENTREGA')

            row += 1

            for j in total_rx_por_recojo:
                hoja.cell(row,1, j.responsable)
                hoja.cell(row,2, j.serial)
                hoja.cell(row,3, j.hora_entrega)
                row +=1

            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="rpt_movimiento_radios_dia_{dia_respaldo}.xlsx"'

            wb.save(response)
            return response
    except Exception as e:
                error_message = f"Error al leer los datos: {e}"
                return HttpResponse(error_message)  
        # return redirect('controlrxevento', id=id)  
@login_required   
def print_view(request, id):

    fecha_hora = d.now()
    # fecha_hora = fecha_hora + timedelta(hours=3)

    return render(request, 'print_seriales.html', {'id':id, 'fecha_hora':fecha_hora})
@login_required
def eliminar_serial(request, id):

    if request.method == 'POST':
        serial_remover = request.POST.get('serial_elim')

        seriales = request.session.get('seriales', [])
        seriales = [item for item in seriales if item.get('serial') != serial_remover]

        request.session['seriales'] = seriales
        request.session.modified = True

        return redirect('controlrxevento', id=id)


@login_required
def mikrotik(request, ip):

    user = 'admin'
    passw = 'B@cktr@ck2OI9'
    listado_wan = []
    listado_ip = []
    listado_interfaz = []
    listado_vlan = []

    try:
        #### BUSCA INFO DHCP - CLIENTES CONECTADOS
        response = requests.get(f'http://{ip}/rest/ip/dhcp-server/lease', auth=HTTPBasicAuth(user,passw),verify=False)
        datos = response.json()

        ##### CUENTA LOS CLIENTES
        count_clientes = 0
        for i in datos:
            count_clientes +=1
        
        listado_ip = [
            {"nombre":item.get("host-name", ""),"address": item["address"]} for item in datos 
            ]
        
        ##### BUSCA INFO DE LAS INTERFACES WAN #################
        response_wan = requests.get(f'http://{ip}/rest/interface', auth=HTTPBasicAuth(user,passw),verify=False)
        datos_wan = response_wan.json()

        #### BUSCAR LA IP DE LA INTERFAZ #####

        for i in datos_wan:
             if i["name"] in ("Wan","Wan1","Wan2","Local","Local 01"):
                response_ip_wan = requests.get(f'http://{ip}/rest/ip/address', auth=HTTPBasicAuth(user,passw),verify=False)
                ip_wan = response_ip_wan.json()
                for j in ip_wan:
                    if i["name"] == j["interface"]:
                        ip_interfaz = j["address"]
                        listado_wan.append({'ip_interfaz':ip_interfaz,'interface':i["name"], 'estado':i["running"]})
        
        ##### OTRAS INTERFACES #################
        response_interfaz = requests.get(f'http://{ip}/rest/interface/bridge/port', auth=HTTPBasicAuth(user,passw),verify=False)
        datos_interfaz = response_interfaz.json()
        listado_interfaz.append({'ip_interfaz_ether':datos_interfaz})

        ############ VLANS ##############

        response_vlan = requests.get(f'http://{ip}/rest/interface/vlan', auth=HTTPBasicAuth(user,passw),verify=False)
        datos_vlan = response_vlan.json()

        for s in datos_vlan:
            vlan_id = s["vlan-id"]
            response_ip_vlan = requests.get(f'http://{ip}/rest/ip/address', auth=HTTPBasicAuth(user,passw),verify=False)
            datos_ip_vlan = response_ip_vlan.json()
            for v in datos_ip_vlan:
                 if s["name"] == v["interface"]:
                    ip_vlan = v["address"]
                    listado_vlan.append({'id_vlan':vlan_id, 'name':s["name"],'interface':s["interface"], 'ip_vlan':ip_vlan, 'activa':s["running"]})
                  

        

        # for k in datos_vlan:
        #           response_ip_vlan = requests.get(f'http://{ip}/ip/address', auth=HTTPBasicAuth(user,passw),verify=False)
        #           datos_ip_vlan = response_ip_vlan.json()
        #           for l in datos_ip_vlan:
        #             if k["name"] == datos_ip_vlan["interface"]:
        #                 listado_vlan.append({'name':k["name"],'interface':k["interface"]})
             

        return render(request, 'mikrotik.html', {'response':listado_ip, 'clientes':count_clientes, 'response_wan':listado_wan, 'ip':ip, 'listado_interfaz_ether':datos_interfaz,
                                                 'listado_vlan':listado_vlan})
    except Exception as e:
                error_message = f"Error al leer los datos: {e}"
                return HttpResponse(error_message)  
@login_required
def cargadash(request):
    user = 'admin'
    passw = 'B@cktr@ck2OI9'
    try:
    # leer la tabla, buscar ip y hacer ping.
        listado_cajas = CajasMikrot.objects.filter(activo = True).order_by('-id')

        result = []

        for i in listado_cajas:
            ip_dir = i.ip
            accesible = False
            respuesta = requests.get(ip_dir, timeout=3)
            if respuesta.status_code == 200:
                accesible = True
                response_nombre = requests.get(f'http://{ip_dir}/rest/system/identity', auth=HTTPBasicAuth(user,passw),verify=False)
                nombre_router = response_nombre.json()
                result.append({'nombre': nombre_router['name'], 'resultado_ping':accesible, 'direccion_ip': ip_dir, 'ubicacion':i.ubicacion})
                #### actualizar el nombre en la tabla 
                # actualiza_nombre = get_object_or_404(CajasMikrot, ip=ip_dir)
                # actualiza_nombre.nombre = nombre_router['name']
                # actualiza_nombre.save()
            else:
                accesible = False
                nombre_router = "DESCONECTADO"
                result.append({'nombre': nombre_router, 'resultado_ping':accesible, 'direccion_ip': ip_dir, 'ubicacion':i.ubicacion})

            
                # buscar el nombre en el json
            

                # result.append({'nombre': nombre_router['name'], 'resultado_ping':accesible, 'direccion_ip': ip_dir, 'ubicacion':i.ubicacion})
        
        return render(request, 'dash_inicial.html',{'resultado':result})
    except Exception as e:
                error_message = f"Error al guardar los datos: {e}"
                return HttpResponse(error_message) 

def get_cajas(request):
    listado_cajas = CajasMikrot.objects.filter(activo=True).order_by('-id')
    result = [{'id': caja.id, 'ip': caja.ip, 'ubicacion': caja.ubicacion} for caja in listado_cajas]
    return JsonResponse({'cajas': result})

def update_router_name(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        ip = data.get('ip')
        nombre = data.get('nombre')

        if ip and nombre:
            caja = get_object_or_404(CajasMikrot, ip=ip)
            caja.nombre = nombre
            caja.save()
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'error': 'Datos incompletos'}, status=400)    



def inventario(request):

    if controlinventario.objects.filter(activo = True).exists():
        fecha_inv_activo = controlinventario.objects.latest('id').fecha_inicio
        id_inv_activo = controlinventario.objects.latest('id').id
        activo = True
        return render(request, 'inventariorx.html', {'activo':activo, 'fecha_inventario':fecha_inv_activo, 'id_inv_activo':id_inv_activo})
    else:
        return render(request, 'inventariorx.html')

def iniciar_inventario(request):

###### APERTURA DEL INVENTARIO ##########
    fecha_inicio = datetime.date.today()
    activo = True

    ## validar si hay otro inventario sin cerrar ###

    if controlinventario.objects.filter(activo = True).exists():
        return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <h2 style="color: red;">YA EXISTE UN PROCESO DE INVENTARIO ACTIVO.</h2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)
    else:
        inicia_inventario = controlinventario(fecha_inicio = fecha_inicio, activo = activo)
        inicia_inventario.save()
    ### BUSCAR ESE ULTIMO ID Y CREAR LA TABLA ESPEJO ########
        id_ult_inventario = controlinventario.objects.latest('id').id
    ##### PREPARAR LA COPIA #####
        datos_inicial = invSeriales.objects.all()
        datos_espejo = []
        for i in datos_inicial:
            dato_espejo = espejo_inventario_ant(id_inventario = id_ult_inventario,
                                                codigo = i.codigo,
                                                estado = i.estado, 
                                                tipo = i.tipo)
            datos_espejo.append(dato_espejo)
        espejo_inventario_ant.objects.bulk_create(datos_espejo)

        ##### renderizar el template con datos #####
        return redirect('inventario')

        
def updaterx_inventario(request, id):

    if request.method == 'POST':
        rx_serial = request.POST.get('serial_rx')

        ### buscarlo en el inv actual #####
        try:
            serial_actual = invSeriales.objects.get(codigo = rx_serial)
            if serial_actual.estado_id != 4:
                ################################################ QUE PASA CUANDO NO ES 2 - PROGRAMAR LOS BOTONES ###########################
                dif_estado = True
                activo = True
                fecha_inv_activo = controlinventario.objects.latest('id').fecha_inicio
                id_inv_activo = controlinventario.objects.latest('id').id
                return render(request, 'inventariorx.html', {'activo':activo, 'fecha_inventario':fecha_inv_activo,
                                                              'id_inv_activo':id_inv_activo, 'dif_estado':dif_estado,
                                                              'rx_serial':rx_serial})
            else:    
                serial_actual.estado_id = 4
                serial_actual.save()
        ### buscarlo en el espejo antes ###
                serial_espejo = espejo_inventario_ant.objects.get(codigo = rx_serial, id_inventario = id)
                serial_espejo.escaneado = True
                serial_espejo.save()
                return redirect('inventario')
        except invSeriales.DoesNotExist:
            return HttpResponse(f"""
                        <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                            <h2 style="color: red;">ESTE SERIAL NO PERTENECE AL INVENTARIO</h2>
                            <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                                Volver atrás
                            </button>
                        </div>
    
                    """)
             
def cambia_estado_dif_inventario(request,estado, serial, id):
    
    serial_actual = invSeriales.objects.get(codigo = serial)

    if estado == 'disponible':
        serial_actual.estado_id = 4
        serial_actual.save()
        ### buscarlo en el espejo antes ###
        serial_espejo = espejo_inventario_ant.objects.get(codigo = serial, id_inventario = id)
        serial_espejo.escaneado = True
        serial_espejo.save()
        return redirect('inventario')
    
    if estado == 'malogrado':
        serial_actual.estado_id = 6
        serial_actual.save()
        ### buscarlo en el espejo antes ###
        serial_espejo = espejo_inventario_ant.objects.get(codigo = serial, id_inventario = id)
        serial_espejo.escaneado = True
        serial_espejo.save()
        return redirect('inventario')


def finaliza_inventario(request, id):

    fecha_cierre = datetime.date.today()
        
    id_ult_inventario = id
    ##### PREPARAR LA COPIA #####
    datos_final = invSeriales.objects.all()
    datos_espejo = []
    for i in datos_final:
        dato_espejo = espejo_inventario_desp(id_inventario = id_ult_inventario,
                                                codigo = i.codigo,
                                                estado = i.estado, 
                                                tipo = i.tipo)
        datos_espejo.append(dato_espejo)
    espejo_inventario_desp.objects.bulk_create(datos_espejo)

    cierra_inventario = controlinventario.objects.get(id = id)
    cierra_inventario.activo = 0
    cierra_inventario.fecha_cierre = fecha_cierre
    cierra_inventario.save() 

    ###### ARMAR EL EXCEL FINAL ###########

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col in sheet.columns:
        col.width = 15.6

    #### GRUPO 1 ########
    gris = PatternFill(start_color='CDCDCD', end_color='CDCDCD', fill_type='solid')
    datos_final = invSeriales.objects.all().order_by('tipo_id', 'codigo')
    
    total_verde_1 = 0
    total_verde_2 = 0
    total_verde_3 = 0

    total_rojo_1 = 0
    total_rojo_2 = 0
    total_rojo_3 = 0

    total_amarillo_1 = 0
    total_amarillo_2 = 0
    total_amarillo_3 = 0

    total_azul_1 = 0
    total_azul_2 = 0
    total_azul_3 = 0

    column = 1
    row = 5
    previo_id = 1

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sheet.cell(1, 1,f"INVENTARIO: {fecha_cierre}").fill = gris
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    # sheet.cell(4, 1,f"EP 450 (Amarillas)").fill = gris
    cell = sheet.cell(4, 1,f"EP 450 (Amarillas)").fill = gris
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for j in datos_final:
        if previo_id != j.tipo_id:
            row += 2
            if j.tipo_id == 2:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                # sheet.cell(row, 1,f"DEP 450 (Moradas)").fill = gris
                cell = sheet.cell(row, 1,f"DEP 450 (Moradas)").fill = gris
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if  j.tipo_id == 3:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                cell = sheet.cell(row, 1,f"DEP 450 (Plomo)").fill = gris
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            column = 1
        if column  == 11:
            sheet.append([])
            row += 1
            column = 1
        verde = PatternFill(start_color='00ff00', end_color='00ff00', fill_type='solid')
        rojo = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
        amarillo = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        azul = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        

        if j.estado_id == 4:
            sheet.cell(row, column,j.codigo).fill = verde
            if j.tipo_id == 1:
                total_verde_1 += 1
            if j.tipo_id == 2:
                total_verde_2 += 1
            if j.tipo_id == 3:
                total_verde_3 += 1
        if j.estado_id == 5:
            sheet.cell(row, column,j.codigo).fill = amarillo
            if j.tipo_id == 1:
                total_amarillo_1 += 1
            if j.tipo_id == 2:
                total_amarillo_2 += 1
            if j.tipo_id == 3:
                total_amarillo_3 += 1
        if j.estado_id == 6:
            sheet.cell(row, column,j.codigo).fill = rojo
            if j.tipo_id == 1:
                total_rojo_1 += 1
            if j.tipo_id == 2:
                total_rojo_2 += 1
            if j.tipo_id == 3:
                total_rojo_3 += 1
        if j.estado_id == 7:
            sheet.cell(row, column,j.codigo).fill = azul
            if j.tipo_id == 1:
                total_azul_1 += 1
            if j.tipo_id == 2:
                total_azul_2 += 1
            if j.tipo_id == 3:
                total_azul_3 += 1

        column += 1
        previo_id = j.tipo_id

    row += 5
    sheet.cell(row, 1,f"Total Disponibles EP 450 (Amarillas): {total_verde_1}")
    row +=1
    sheet.cell(row, 1,f"Total Disponibles DEP 450 (Moradas): {total_verde_2}")
    row +=1
    sheet.cell(row, 1,f"Total Disponibles DEP 450 (Plomo): {total_verde_3}")
    row +=2
    sheet.cell(row, 1,f"Total Malogradas EP 450 (Amarillas): {total_rojo_1}")
    row +=1
    sheet.cell(row, 1,f"Total Malogradas DEP 450 (Moradas): {total_rojo_2}")
    row +=1
    sheet.cell(row, 1,f"Total Malogradas DEP 450 (Plomo): {total_rojo_3}")
    row +=2
    sheet.cell(row, 1,f"Total No Disponibles EP 450 (Amarillas): {total_amarillo_1}")
    row +=1
    sheet.cell(row, 1,f"Total No Disponibles DEP 450 (Moradas): {total_amarillo_2}")
    row +=1
    sheet.cell(row, 1,f"Total No Disponible DEP 450 (Plomo): {total_amarillo_3}")
    row +=2
    sheet.cell(row, 1,f"Total Extraviadas EP 450 (Amarillas): {total_azul_1}")
    row +=1
    sheet.cell(row, 1,f"Total Extraviadas DEP 450 (Moradas): {total_azul_2}")
    row +=1
    sheet.cell(row, 1,f"Total Extraviadas DEP 450 (Plomo): {total_azul_3}")
    
    ##### LEYENDA #####
    row +=3
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sheet.cell(row, 1,f"Disponibles")
    sheet.cell(row, 3,f" ").fill = verde
    row += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sheet.cell(row, 1,f"No Disponibles")
    sheet.cell(row, 3,f" ").fill = amarillo
    row += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sheet.cell(row, 1,f"Malogradas")
    sheet.cell(row, 3,f" ").fill = rojo
    row += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    sheet.cell(row, 1,f"Extraviadas")
    sheet.cell(row, 3,f" ").fill = azul


    response = HttpResponse(content_type = 'application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Inventario_{fecha_cierre}.xlsx"'

    workbook.save(response)
    return response

    
def inventario_acce(request):

    if controlinventarioacce.objects.filter(activo = True).exists():
        fecha_inv_activo = controlinventarioacce.objects.latest('id').fecha_inicio
        id_inv_activo = controlinventarioacce.objects.latest('id').id
        activo = True
        item = inv_accesorios_temp.objects.all().order_by('id_item')
        return render(request, 'inventarioacce.html', {'activo':activo, 'fecha_inventario':fecha_inv_activo, 'id_inv_activo':id_inv_activo, 'item':item})
    else:
        return render(request, 'inventarioacce.html')
    # return render(request, 'inventarioacce.html')
 
def iniciainventarioacce(request):
    fecha_inicio = datetime.date.today()
    activo = True

    if controlinventarioacce.objects.filter(activo = True).exists():
        return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <h2 style="color: red;">YA EXISTE UN PROCESO DE INVENTARIO ACTIVO.</h2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)
    else:
        iniciainventarioacce = controlinventarioacce(fecha_inicio = fecha_inicio, activo = activo)
        iniciainventarioacce.save()

        id_ult_inventario = controlinventarioacce.objects.latest('id').id

        datos_inicial = inv_accesorios.objects.all().order_by('id')
        datos_espejo = []
        datos_temp = []
        for i in datos_inicial:
            dato_espejo = espejo_inventarioacce_ant(id_inventario = id_ult_inventario,
                                                id_item = i.id,
                                                cantidad = i.cantidad)
            dato_temp = inv_accesorios_temp(id_item = i.id, descripcion = i.descripcion, cantidad = 0)
            datos_temp.append(dato_temp)
            datos_espejo.append(dato_espejo)
        
        espejo_inventarioacce_ant.objects.bulk_create(datos_espejo)
        inv_accesorios_temp.objects.bulk_create(datos_temp)

        return redirect('inventario_acce')

def updateinventarioacce(request, id):
    
    if request.method == 'POST':
        cantidad = request.POST.get('cantid')
        nueva_cantidad = inv_accesorios_temp.objects.get(id_item = id)
        nueva_cantidad.cantidad = nueva_cantidad.cantidad + int(cantidad)
        nueva_cantidad.save()

        return redirect('inventario_acce')    

def finaliza_inventarioacce(request, id):

    fecha_cierre = datetime.date.today()
        
    id_ult_inventario = id
    ##### PREPARAR LA COPIA #####
    datos_final = inv_accesorios_temp.objects.all().order_by('id')
    datos_espejo = []
    for i in datos_final:
        dato_espejo = espejo_inventarioacce_desp(id_inventario = id_ult_inventario,
                                                id_item = i.id_item,
                                                cantidad = i.cantidad)
        datos_espejo.append(dato_espejo)

        actu_final = inv_accesorios.objects.get(id = i.id_item)
        actu_final.cantidad = i.cantidad
        actu_final.save()
        


    espejo_inventarioacce_desp.objects.bulk_create(datos_espejo)

    ### CERRAR EL INVENTARIO #####

    cierra_inventario = controlinventarioacce.objects.get(id = id)
    cierra_inventario.activo = 0
    cierra_inventario.fecha_cierre = fecha_cierre
    cierra_inventario.save()


    ##### BORRAR TEMPORAL AL CERRAR #####

    inv_accesorios_temp.objects.all().delete()

    ### PARA EL REPORTE FINAL ####

    # con el id_ultmo - 1 buscar la fecha de cierre anterior 

    # id_inv_anterior = id_ult_inventario -1
    # fecha_anterior = controlinventarioacce.objects.get(id = id_inv_anterior)
    # fecha_cierre_antes = fecha_anterior.fecha_cierre
    
    return redirect('inventario_acce') 

def cargar_entrada_accesorios(request):

    item = inv_accesorios.objects.all().order_by('id')

    return render(request, 'entrada_accesorios.html',{'item':item})

def actualiza_entrada_acce(request):

    fecha_entrada = datetime.date.today()

    if request.method == 'POST':
        
        if controlinventarioacce.objects.filter(activo = True).exists():
            return HttpResponse(f"""
                    <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                    <h2 style="color: red;">EXISTE UN PROCESO DE INVENTARIO ACTIVO.</h2>
                    <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                    Volver atrás
                    </button>
                    </div>
                    """)
        else:
            id = request.POST.get('itemaccesorio')
            cantidad = int(request.POST.get('cantid'))

            # CREAR EL REGISTRO DE ENTRADA #

            crea_registro = entrada_accesorios(id_item = id, cantidad = cantidad, fecha_entrada = fecha_entrada)
            crea_registro.save()

            # ACTUALIZA EL INVENTARIO #

            item_actualizar = inv_accesorios.objects.get(id = id)
            item_actualizar.cantidad = item_actualizar.cantidad + cantidad
            item_actualizar.save()

            messages.success(request, '¡EL INVENTARIO HA SIDO ACTUALIZADO!')

            return redirect('cargar_entrada_accesorios')    

def consulta_inventario_acce(request):

    item = inv_accesorios.objects.all().order_by('id')

    return render(request, 'consulta_inv_accesorios.html', {'item':item})    

def kardex(request):

    rpt_kardex.objects.all().delete()

    try:
        ult_inv = controlinventarioacce.objects.filter(activo=False).latest('fecha_cierre')
        fecha_ult_inv = ult_inv.fecha_cierre
        id_inventario = ult_inv.id
    except controlinventarioacce.DoesNotExist:
    
        fecha_ult_inv = None
        id_inventario = None

    if fecha_ult_inv and id_inventario:
        existencia_act = inv_accesorios.objects.all()

    for i in existencia_act:
        id_item = i.id
        descripcion_item = i.descripcion
        cant_existencia = i.cantidad or 0

        # Última actualización
        ult_actualizacion = espejo_inventarioacce_desp.objects.filter(
            id_inventario=id_inventario, id_item=id_item
        ).first()
        ult_actualizacion_cant = ult_actualizacion.cantidad if ult_actualizacion else 0

        # Total entradas
        total_entradas = entrada_accesorios.objects.filter(
            fecha_entrada__gte=fecha_ult_inv, id_item=id_item
        ).aggregate(total=Sum('cantidad'))['total'] or 0

        # Total salidas
        total_salidas = entrada_salida_acce.objects.filter(
            fecha_mov__gte=fecha_ult_inv, tipo_mov='S', id_item=id_item
        ).aggregate(total=Sum('cantidad'))['total'] or 0

        # Entradas agrupadas
        total_entradas_acce = entrada_salida_acce.objects.filter(
            fecha_mov__gte=fecha_ult_inv, tipo_mov='E', id_item=id_item
        ).aggregate(total=Sum('cantidad'))['total'] or 0

        # Calcular diferencia
        suma_alge = (
            ult_actualizacion_cant + total_entradas + total_entradas_acce - total_salidas
        )
        diferencia = cant_existencia - suma_alge
    
        guardar_kardex = rpt_kardex(id_item = id_item, descripcion = descripcion_item, ult_actualizacion = ult_actualizacion_cant,
                                     entrada_merc =  total_entradas, salidas = total_salidas, entradas = total_entradas_acce,
                                     existencia_actual = cant_existencia, dif = diferencia)
        guardar_kardex.save()

    datos = rpt_kardex.objects.all()    
    return render(request, 'kardex.html', {'datos':datos})
        
      
        
def exportar_radios_excel(request):


    inv_disp = invSeriales.objects.filter(estado_id =4)
    cuenta_disponibles = invSeriales.objects.filter(estado_id =4).count()
    inv_nodisp = invSeriales.objects.filter(estado_id =5)
    cuenta_nodisponibles = invSeriales.objects.filter(estado_id =5).count()
    inv_malogrado = invSeriales.objects.filter(estado_id =6)
    cuenta_malogrado = invSeriales.objects.filter(estado_id =6).count()
    amarillas_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 1).count()
    moradas_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 2).count()
    plomo_disp = invSeriales.objects.filter(estado_id =4, tipo_id = 3).count()
    amarillas_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 1).count()
    moradas_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 2).count()
    plomo_nodisp = invSeriales.objects.filter(estado_id =5, tipo_id = 3).count()
    extraviadas = invSeriales.objects.filter(estado_id =7).count()
    total_radios = invSeriales.objects.filter().count()

    
    
    #crea nuevo libro
    wb = openpyxl.Workbook()

    hoja = wb.active

    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    hoja['A1'].alignment = Alignment(horizontal="center", vertical="center")
   
    hoja['A1'] = 'RELACIÓN DE RADIOS'
    hoja['A1'].fill = fill

    hoja['A3'] = 'DISPONIBLES'
    hoja['A4'] = 'AMARILLAS'
    hoja['B4'] = amarillas_disp
    hoja['A5'] = 'MORADAS'
    hoja['B5'] = moradas_disp
    hoja['A6'] = 'PLOMO'
    hoja['B6'] = plomo_disp
    hoja['A7'] = 'TOTAL'
    hoja['B7'] = cuenta_disponibles

    hoja['A9'] = 'NO DISPONIBLES'
    hoja['A10'] = 'AMARILLAS'
    hoja['B10'] = amarillas_nodisp
    hoja['A11'] = 'MORADAS'
    hoja['B11'] = moradas_nodisp
    hoja['A12'] = 'PLOMO'
    hoja['B12'] = plomo_nodisp
    hoja['A13'] = 'TOTAL'
    hoja['B13'] = cuenta_nodisponibles

    hoja['A15'] = 'MALOGRADAS'
    hoja['B15'] = cuenta_malogrado

    hoja['A17'] = 'EXTRAVIADAS'
    hoja['B17'] = extraviadas

    hoja['A19'] = 'TOTAL RADIOS EN INVENTARIO'
    hoja['B19'] = total_radios

    hoja.column_dimensions['A'].width = 30
    hoja.column_dimensions['B'].width = 15



 
    response = HttpResponse(content_type = 'application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="relacion_radios.xlsx"'

    wb.save(response)

    return response

 
def entradasfalt(request, id,serial_rx):

    form = guardaEntradaRx()
    formEntrada = formEntradaDetalle()
    context = {'form':form}

    serial = serial_rx
    msalida = salidasDetalle.objects.get(id=id)
    orden_id = msalida.id_orden
    ordenes = ordenRegistro.objects.get(id = orden_id)

#buscar las radios cargadas

    radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")
    return render(request, 'entradarxfaltante.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada, 'serial':serial})   


def entradasfaltanterx(request, id, orden_id, serial_rx):

    form = guardaEntradaRx()
    formEntrada = formEntradaDetalle()
    #verificar aqui el id
    radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")

    u = orden_id

    msalida = salidasDetalle.objects.get(id=id)
    orden_id = msalida.id_orden
    ordenes = ordenRegistro.objects.get(id = orden_id)


    if request.method == 'POST':

        form = guardaEntradaRx(request.POST)
        
        if form.is_valid():
            a = form.cleaned_data['serial']

            if a != serial_rx:
                return HttpResponse(f"""
                <div style="display: flex; justify-content: center; align-items: center; height: 100vh; flex-direction: column; background-color: #f0f0f0; padding: 20px; border-radius: 5px; font-family: Arial, sans-serif; font-style: italic;">
                <img src="/static/error.png" alt="Error" style="max-width: 200px; margin-bottom: 20px;">
                <h2 style="color: red;">EL SERIAL INTRODUCIDO NO CORRESPONDE</h2>
                <button onclick="history.back()" style="padding: 10px 20px; background-color: #f0ad4e; border: none; border-radius: 5px; color: white; cursor: pointer;">
                Volver atrás
                </button>
                </div>
                """)
            ##comprobar que existe una salida con ese serial y ese numero de salida
            if not movimientoRadios.objects.filter(serial = a, id_salida = id, estado = "F").exists():
                initial_data = {
                    'serial': '',
                    }
                form = guardaEntradaRx(initial=initial_data)
                return render(request, 'entradarxfaltante.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                                        "error": "El serial ingresado no corresponde a la salida", "formEntrada":formEntrada})
            ##valida si ya fue agregado
            if movimientoRadios.objects.filter(serial = a, id_salida = id, estado = "D").exists():
                initial_data = {
                    'serial': '',
                    }
                form = guardaEntradaRx(initial=initial_data)
                return render(request, 'entradarxfaltante.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                                        "error": "El serial ingresado ya fue agregado", "formEntrada":formEntrada})
            
            else:
                valExisteRxSalida = movimientoRadios.objects.get(serial = a, id_salida = id, estado = "F")
                a = form.save(commit=False)
                a.id_salida = id
                a.estado = 'D'
                a.id_tipo = valExisteRxSalida.id_tipo
                a.save()

                d = a.serial

                #borrar si existe en faltantes

                if radiosFantantes.objects.filter(fserial = d, id_salida = id).exists():
                    borrarrx = radiosFantantes.objects.filter(fserial = d, id_salida = id)
                    borrarrx.delete()


                #cambiar el estatus del radio en el inventario

                cambiaestado =  invSeriales.objects.get(codigo = d)
                cambiaestado.estado_id = 4
                cambiaestado.save()

                ## afectar inventario accesorios

                # BATERIAS

                t_suma_bat = inv_accesorios.objects.get(id = 5)
                t_mas_baterias = 1
                t_suma_bat.cantidad = t_suma_bat.cantidad + t_mas_baterias
                t_suma_bat.save()

                fecha_mov = datetime.date.today()

                f_suma_bat = entrada_salida_acce(id_item = 5, cantidad = t_mas_baterias,
                                                                  tipo_mov = 'E', fecha_mov = fecha_mov)
                f_suma_bat.save()

                


            ordenes = ordenRegistro.objects.get(id = orden_id)
            msalida = salidasDetalle.objects.get(id_orden=orden_id)

            initial_data = {
                    'serial': '',
                    }
            form = guardaEntradaRx(initial=initial_data)

            radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")
            return redirect('/verFaltante/') 

    return render(request,'entradarxfaltante.html', {"form":form})

    
def entradaaccefaltante(request, id):

    form = guardaEntradaRx()
    formEntrada = formEntradaDetalle()
    context = {'form':form}

    
    msalida = salidasDetalle.objects.get(id=id)
    orden_id = msalida.id_orden
    ordenes = ordenRegistro.objects.get(id = orden_id)

#buscar las radios cargadas

    radiosCargadas = movimientoRadios.objects.filter(id_salida = id, estado = "D")
    return render(request, 'entradaaccefaltante.html',{"listado_entrada": msalida, "listado_orden":ordenes, "listadoRadiosCargadas":radiosCargadas, "form":form, 
                                             "formEntrada":formEntrada})


def carga_formulario_pedido(request):

    try:
        form = FormPedidoCliente()

        if request.method == 'POST':
            form = FormPedidoCliente(request.POST)
            if form.is_valid():
                f_creacion = datetime.date.today()
                nombre = form.cleaned_data['Nombre']
                telefono = form.cleaned_data['telefono']
                f_entrega = form.cleaned_data['fecha_entrega']
                f_evento = form.cleaned_data['fecha_evento']
                cantidad_rx = form.cleaned_data['cantidad_radios']
                cantidad_cobras = form.cleaned_data['cantidad_cobras']
                cantidad_manos_l = form.cleaned_data['cantidad_manos_libres']
                cantidad_t_escolta = form.cleaned_data['cantidad_tipo_escolta']
                direccion = form.cleaned_data['direccion_entrega']
                comentarios = form.cleaned_data['comentarios']

                guardar_form = form.save(commit=False)
                guardar_form.fecha_creacion = f_creacion
                guardar_form.save()

                ############ ENVIARLO POR WHATSAPP ############

                # telwhat = +51946364693
                telwhat = ["+51933805151", "+51974616099", "+51946364693","+51914142952"] 
                mensaje = f'''Hola. Tienes un nuevo pedido.
Cliente: {nombre}
Teléfono: {telefono}
Dirección de Entrega: {direccion}
Comentarios del Cliente: {comentarios}\n
*DATOS DEL PEDIDO:*
Fecha de Entrega: {f_entrega}
Fecha del Evento: {f_evento}
Radios: {cantidad_rx}
Cobras: {cantidad_cobras}
Manos Libres: {cantidad_manos_l}
Tipo Escolta: {cantidad_t_escolta}'''
                
                mensaje_codificado = urllib.parse.quote(mensaje)
                for tel in telwhat:
                    url = "https://api.ultramsg.com/instance109145/messages/chat"
                    payload = f"token=k7wzcqez2zfna7e2&to=%2B{tel}&body={mensaje_codificado}"
                    # payload = payload.encode('utf8').decode('iso-8859-1')
                    headers = {'content-type': 'application/x-www-form-urlencoded'}
                    response = requests.request("POST", url, data=payload, headers=headers)

                mensaje_cliente = f'''Estimado/a {nombre},
Hemos recibido su pedido. 
En breve, un miembro de nuestro equipo se pondrá en contacto con usted para brindarle más detalles y coordinar cualquier información adicional.
Si tiene alguna consulta, no dude en escribirnos. Estamos aquí para ayudarle.\n
¡Gracias por confiar en nosotros!\n
Atte.
*BACKTRAK SOLUTION NETWORK*'''
                if not telefono is None:
                    telefono = re.sub(r'[^\d]+', '', telefono)
                    telefono = "51" + telefono if not telefono.startswith("51") else telefono

                mensaje_cod = urllib.parse.quote(mensaje_cliente)

                url = "https://api.ultramsg.com/instance109145/messages/chat"
                payload = f"token=k7wzcqez2zfna7e2&to=%2B{telefono}&body={mensaje_cod}"
                # payload = payload.encode('utf8').decode('iso-8859-1')
                headers = {'content-type': 'application/x-www-form-urlencoded'}
                response = requests.request("POST", url, data=payload, headers=headers)

                return redirect('gracias')
        return render(request, 'formulario_pedido.html',{'form':form})
    except Exception as e:
                error_message = f"Error al guardar los datos: {e}"
                return HttpResponse(error_message) 


def gracias(request):

    return render(request, 'gracias.html')

def enviar_recordatorios(request):

    fecha_hoy = datetime.date.today()

    por_entregar = ordenRegistro.objects.filter(fecha_entrega = fecha_hoy, estado_id = 5)
    por_recoger = ordenRegistro.objects.filter(fecha_retiro = fecha_hoy, estado_id = 3)

    return render(request, 'enviar_recordatorio.html',{'por_entregar':por_entregar, 'por_recoger':por_recoger})

def envio_whatsapp_entrega(request):

    fecha_hoy = datetime.date.today()

    por_entregar = ordenRegistro.objects.filter(fecha_entrega = fecha_hoy, estado_id = 5)
    
    if request.method == "POST":
        telwhats = ["+51933805151", "+51914142952", "+51946364693","+51918961967"] 
        mensaje = "*PEDIDOS POR ENTREGAR HOY:*\n"

        if por_entregar.exists():
            for orden in por_entregar:
                if not orden.direccion_entrega:
                    direccion = 'Por confirmar'
                else: 
                    direccion = orden.direccion_entrega
                mensaje += f'''
*Cliente:* {orden.cliente}
*Dirección:* {direccion}\n'''
                
            mensaje_codificado = urllib.parse.quote(mensaje)

            for tel in telwhats:
                url = "https://api.ultramsg.com/instance109145/messages/chat"
                payload = f"token=k7wzcqez2zfna7e2&to=%2B{tel}&body={mensaje_codificado}"
                headers = {'content-type': 'application/x-www-form-urlencoded'}
                response = requests.request("POST", url, data=payload, headers=headers)

        return redirect('recordatorio')
    return redirect('recordatorio')

def envio_whatsapp_recojo(request):

    fecha_hoy = datetime.date.today()
    por_recoger = ordenRegistro.objects.filter(fecha_retiro = fecha_hoy, estado_id = 3)

    if request.method == "POST":
        telwhats = ["+51933805151", "+51914142952", "+51946364693","+51918961967"] 
        mensaje = "*PEDIDOS POR RECOGER HOY:*\n"

        if por_recoger.exists():
            for orden in por_recoger:
                if not orden.direccion_entrega:
                    direccion = 'Por confirmar'
                else: 
                    direccion = orden.direccion_entrega
                mensaje += f'''
*Cliente:* {orden.cliente}
*Dirección:* {direccion}\n'''
            
            mensaje_codificado = urllib.parse.quote(mensaje)
            # url = "https://api.ultramsg.com/instance108729/messages/chat"
            for tel in telwhats:
                url = "https://api.ultramsg.com/instance109145/messages/chat"
                payload = f"token=k7wzcqez2zfna7e2&to=%2B{tel}&body={mensaje_codificado}"
                headers = {'content-type': 'application/x-www-form-urlencoded'}
                response = requests.request("POST", url, data=payload, headers=headers)

        return redirect('recordatorio')
    return redirect('recordatorio')

def consultar_estado_pedido(request):
    numero_pedido = request.GET.get("numero_pedido")  # Obtener el número del pedido desde la URL

    try:
        numero_pedido = int(numero_pedido)  # Convertir a número
    except ValueError:
        return JsonResponse({"error": "Número de pedido inválido"}, status=400)

    if not numero_pedido:
        return JsonResponse({"error": "Número de pedido requerido"}, status=400)

    pedido = ordenRegistro.objects.get(id=numero_pedido)

    if pedido.estado_id == 1:
        estado_texto = "Anulado"
    elif pedido.estado_id == 2:
        estado_texto = "Pedido"
    elif pedido.estado_id == 3:
        estado_texto = "Entregado"
    elif pedido.estado_id == 4:
        estado_texto = "Devuelto"
    elif pedido.estado_id == 5:
        estado_texto = "Devuelto"
    else:
        estado_texto = "Desconocido"

    response_data = { "data": { "type": estado_texto, "contacts" : 
                               [ { "firstName":"Steve", "lastName":"De", "email": "steve@example.com" }, 
                                { "firstName":"Jane", "lastName":"Doe", "email": "Jane@example.com" } ] } }
    return JsonResponse(response_data)
        

def enviar_mensaje(request, id):

    ####### BUSCAR LAS PERSONAS CON TELEFONO Y ESTATUS E #######
    ####### REVISAR SI TIENE VARIAS RADIOS ###### SOLO SE DEBE ENVIAR UN MENSAJE ###########
    personas_envio_queryset = controlrxevent.objects.filter(id_salida=id, estadorx = 'E').exclude(Q(telefono__isnull=True) | Q(telefono=''))

    personas_envio_dict = {}
    for persona in personas_envio_queryset:
        if persona.telefono not in personas_envio_dict:
            personas_envio_dict[persona.telefono] = persona

    personas_envio = list(personas_envio_dict.values())

    for i in personas_envio:
        telwhat = i.telefono
        if not telwhat is None:
                telwhat = re.sub(r'[^\d]+', '', telwhat)
                telwhat = "51" + telwhat if not telwhat.startswith("51") else telwhat
        
        mensaje = """Buenas.
Le recordamos que las radios en su poder deben ser devueltas al finalizar el evento.
Agradecemos su colaboración.
Quedamos atentos para coordinar la entrega.\n
Atte.
*BACKTRAK SOLUTION NETWORK*"""
        mensaje_codificado = urllib.parse.quote(mensaje)

        url = "https://api.ultramsg.com/instance109145/messages/chat"
        payload = f"token=k7wzcqez2zfna7e2&to=%2B{telwhat}&body={mensaje_codificado}"
        headers = {'content-type': 'application/x-www-form-urlencoded'}
        response = requests.request("POST", url, data=payload, headers=headers)

    return redirect('controlrxevento', id=id) 

def asignar_pedido_preparado(request):

    #verificar si ya existe
    if request.method == 'POST':
        id_orden = request.POST.get('orden_id')
        usuario_asignado = request.POST.get('itemasignado')
        fecha_hoy = datetime.date.today()
        hora_actual = d.now()
        # hora_actual = hora_actual + timedelta(hours=3)
        hora_actual = hora_actual.strftime("%H:%M")

        if not pedidos_asignados_prepa.objects.filter(id_orden = id_orden).exists():
            pedidos_asignados_prepa.objects.create(id_orden = id_orden, asignado_a = usuario_asignado, fecha_asignacion = fecha_hoy, hora_asignacion = hora_actual)
            #### ENVIAR MENSAJE
            return redirect('ordenesProcesadas')
        else:
            pedidos_asignados_prepa.objects.filter(id_orden=id_orden).update(asignado_a=usuario_asignado)
            #### ENVIAR MENSAJE
            return redirect('ordenesProcesadas')


def asignar_pedido_entregado(request):

    #verificar si ya existe
    if request.method == 'POST':
        id_orden = request.POST.get('orden_id')
        usuario_asignado = request.POST.get('itemasignado')
        fecha_hoy = datetime.date.today()
        hora_actual = d.now()
        # hora_actual = hora_actual + timedelta(hours=3)
        hora_actual = hora_actual.strftime("%H:%M")

        if not pedidos_asignados_entrega.objects.filter(id_orden = id_orden).exists():
            pedidos_asignados_entrega.objects.create(id_orden = id_orden, asignado_a = usuario_asignado, fecha_asignacion = fecha_hoy, hora_asignacion = hora_actual)
            #### ENVIAR MENSAJE
            return redirect('ordenesCerradas')
        else:
            pedidos_asignados_entrega.objects.filter(id_orden=id_orden).update(asignado_a=usuario_asignado)
            #### ENVIAR MENSAJE
            return redirect('ordenesCerradas')